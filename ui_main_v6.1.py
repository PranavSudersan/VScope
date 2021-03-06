# -*- coding: utf-8 -*-
"""
Created on Sun May 26 23:28:45 2019

@author: pranav
"""
# %% Import modules
import cv2
import numpy as np
import matplotlib.pyplot as plt
import time
import sys
import os.path
import draw_roi
from videoTransform import Effects
from analyze_force import ForceAnal
from summarize import SummaryAnal
from PyQt5.QtGui import QPixmap, QIcon, QImage, QBrush, \
     QPolygonF, QPainter, QPalette, QColor
from PyQt5.QtCore import Qt, QThread, pyqtSignal, QFile, QTextStream
from PyQt5.QtWidgets import QApplication, QMainWindow, \
     QSlider, QFileDialog, QCheckBox, QLabel, QPushButton, \
     QMessageBox, QAction, QComboBox, QGraphicsScene, QGraphicsView, \
     QGraphicsPixmapItem, QLineEdit, QTextEdit, QSpinBox, QDoubleSpinBox, QStatusBar, \
     QGroupBox, QGridLayout, QVBoxLayout, QHBoxLayout, QWidget, QSizePolicy, \
     QDialog, QListWidget
from PyQt5.QtChart import QChart, QScatterSeries
from tkinter import messagebox, Tk
import pandas as pd
from PIL import ImageFont, ImageDraw, Image
import openpyxl
import copy
from cv2_rolling_ball import subtract_background_rolling_ball
import gc
##import qdarkstyle
##from style import breeze_resources

# %% Main Application Window
class MainWindow(QMainWindow, Effects): #also try inherit Effect, unify self.frame everywhere
    def __init__(self):
        super().__init__()
        self.setGeometry(0, 0, 1500, 1100)
        self.appVersion = "V-Scope v6.1"
        self.setWindowTitle(self.appVersion)
        self.layout = QGridLayout()
        
        self.configROIWindow = ConfigROIWindow() #ROI config
        self.configPathWindow = ConfigPathWindow() #path window
        self.configRecWindow = ConfigRecWindow() #record configuration window
        self.configPlotWindow = ConfigPlotWindow() #plot configuration window
        self.plotWindow = PlotWindow() #live plot window

        quitWindow = QAction("&Quit", self) #quit program
##        quitWindow.setShortcut("Ctrl+Q") 
        quitWindow.setStatusTip('Quit Program')
        quitWindow.triggered.connect(self.closeEvent)
        
        openVideoFile = QAction("&Open Video File...", self) #open video file
##        openVideoFile.setShortcut("Ctrl+V")
        openVideoFile.setStatusTip("Select the required video file")
        openVideoFile.triggered.connect(self.load_video)

        openImageFile = QAction("&Open Image File...", self) #open image file
##        openImageFile.setShortcut("Ctrl+I")
        openImageFile.setStatusTip("Select the required image file")
        openImageFile.triggered.connect(self.load_image)

        openForceFile = QAction("&Open Force Data...", self) #open force file
##        openForceFile.setShortcut("Ctrl+F")
        openForceFile.setStatusTip("Select the corresponding force data file")
        openForceFile.triggered.connect(self.import_force_data)

        self.openZeroForceFile = QAction("&Open Zero-Force Data...", self) #open zero force file
        self.openZeroForceFile.setStatusTip("Select the zeroline force data file")
        self.openZeroForceFile.triggered.connect(self.import_zero_force)

        openFileList = QAction("&Open File List...", self) #open file list spreadhsheet
##        openFileList.setShortcut("Ctrl+O")
        openFileList.setStatusTip("Select file list spreadhsheet")
        openFileList.triggered.connect(self.import_file_list)

        self.chooseMsrmnt = QAction("&Choose Measurement...", self) #choose measurement from file list
##        self.chooseMsrmnt.setShortcut("Ctrl+M")
        self.chooseMsrmnt.setStatusTip("Select measurement from list")
        self.chooseMsrmnt.triggered.connect(self.choose_measurement)

        recordVideo = QAction("&Record", self) #configure recording
##        recordVideo.setShortcut("Ctrl+R")
        recordVideo.setStatusTip("Configure Record")
        recordVideo.triggered.connect(self.showRecWindow)
        self.configRecWindow.okBtn.clicked.connect(self.configureRecord)

        plot = QAction("&Plot/Force", self) #configure plot
##        plot.setShortcut("Ctrl+P")
        plot.setStatusTip("Configure plot and force curves")
        plot.triggered.connect(self.configPlotWindow.show_window)
        self.configPlotWindow.okBtn.clicked.connect(self.configurePlot)
        self.configPlotWindow.updateBtn.clicked.connect(self.plotSequence)

        paths = QAction("&Filepath", self) #configure export file paths
        paths.setStatusTip("Configure Filepaths")
        paths.triggered.connect(self.showPathWindow)
        self.configPathWindow.okBtn.clicked.connect(self.setPaths)

        showSummary = QAction("&Display Summary Plots", self) #show summary Plots
        showSummary.setStatusTip("Displays summary plots based on summary data file")
        showSummary.triggered.connect(lambda: self.show_summary_plots())

        exportSummary = QAction("&Export Summary Plots", self) #export summary Plots
        exportSummary.setStatusTip("Exports summary plots based on summary data file")
        exportSummary.triggered.connect(self.export_summary_plots)

##        combineSummary = QAction("&Combine Summary Data", self) #combine data from list
##        combineSummary.setStatusTip("Combines summary data from list of file names")
##        combineSummary.triggered.connect(self.summary_dialog)
##        self.sumDialog = QDialog(self) #summary dialog
##        self.sumDialog.setWindowTitle("Configure Summary Plots")
        self.summary_dialog_init()
        configureSummary = QAction("&Configure", self) #configure summary plots
        configureSummary.setStatusTip("Configure Summary Plot")
        configureSummary.triggered.connect(lambda: self.sumDialog.show())
        
        mainMenu = self.menuBar() #create main menu
        
        fileMenu = mainMenu.addMenu("&File") #File menu
        fileMenu.addAction(openVideoFile)
        fileMenu.addAction(openImageFile)
        fileMenu.addAction(openForceFile)
        fileMenu.addAction(self.openZeroForceFile)
        fileMenu.addAction(openFileList)
        fileMenu.addAction(self.chooseMsrmnt)
        fileMenu.addAction(quitWindow)

        self.chooseMsrmnt.setEnabled(False)
        self.openZeroForceFile.setEnabled(False)

        configureMenu = mainMenu.addMenu("&Configure") #Configure menu
        configureMenu.addAction(recordVideo)
        configureMenu.addAction(plot)
        configureMenu.addAction(paths)

        plotMenu = mainMenu.addMenu("&Summarize") #Plot menu
        plotMenu.addAction(configureSummary)
        plotMenu.addAction(showSummary)
        plotMenu.addAction(exportSummary)
##        plotMenu.addAction(combineSummary)

        self.statusBar = QStatusBar() #status bar
        self.setStatusBar(self.statusBar)
        
        self.home()
        

        
    def home(self):

        self.blankPixmap = QPixmap('images/blank.png')
        
        
        self.rawScene = QGraphicsScene(self) #raw video feed
        self.rawPixmapItem = QGraphicsPixmapItem(self.blankPixmap)
        self.rawScene.addItem(self.rawPixmapItem)
        self.rawView = MyQGraphicsView(self.rawScene)
##        self.rawView.setGeometry(QRect(50, 120, 480, 360))
        self.rawView.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.rawView.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.rawView.setBackgroundBrush(QBrush(Qt.black,
                                                       Qt.SolidPattern))

        self.effectScene = QGraphicsScene(self) #analysis video feed
##        self.effectPixmap = QPixmap('images/blank.png')
        self.effectPixmapItem = QGraphicsPixmapItem(self.blankPixmap)
        self.effectScene.addItem(self.effectPixmapItem)
        self.effectView = MyQGraphicsView(self.effectScene)
##        self.effectView.setGeometry(QRect(560, 120, 480, 360))
        self.effectView.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.effectView.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.effectView.setBackgroundBrush(QBrush(Qt.black,
                                                          Qt.SolidPattern))
        
        self.playBtn = QPushButton("", self) #Play/Pause action
        self.playBtn.clicked.connect(self.playback)
        self.playBtn.setIcon(QIcon('images/play.png'))
##        self.playBtn.resize(self.playBtn.minimumSizeHint())
##        self.playBtn.move(100,500)

        self.stopBtn = QPushButton("", self) #Stop video
        self.stopBtn.clicked.connect(self.stop_video)
        self.stopBtn.setIcon(QIcon('images/stop.png'))
##        self.stopBtn.resize(self.stopBtn.minimumSizeHint())
##        self.stopBtn.move(150,500)

        self.prevBtn = QPushButton("", self) #Previous frame
        self.prevBtn.clicked.connect(self.previous_frame)
        self.prevBtn.setIcon(QIcon('images/previous.png'))
##        self.prevBtn.resize(self.prevBtn.minimumSizeHint())
##        self.prevBtn.move(200,500)

        self.nextBtn = QPushButton("", self) #Next frame
        self.nextBtn.clicked.connect(self.next_frame)
        self.nextBtn.setIcon(QIcon('images/next.png'))
##        self.nextBtn.resize(self.nextBtn.minimumSizeHint())
##        self.nextBtn.move(250,500)

        self.recordBtn = QPushButton("", self) #Record video
        self.recordBtn.clicked.connect(self.record_frame)
        self.recordBtn.setIcon(QIcon('images/record.png'))
##        self.recordBtn.resize(self.recordBtn.minimumSizeHint())
##        self.recordBtn.move(300,500)
        self.recordBtn.setEnabled(False)
        
        self.seekSlider = QSlider(Qt.Horizontal, self) #seek bar
##        self.seekSlider.setGeometry(30, 60, 200, 30)
        self.seekSlider.setMinimum(0)
        self.seekSlider.setMaximum(100)
        self.seekSlider.setValue(0)
        self.seekSlider.setTickInterval(10)
        self.seekSlider.setSingleStep(2)
        self.seekSlider.setTickPosition(QSlider.TicksBelow)
        self.seekSlider.valueChanged.connect(self.seek_video)
        
        self.roiBtn = QPushButton("Draw ROI", self) #Draw ROI Manually
        self.roiBtn.clicked.connect(self.configROIWindow.showWindow)
        self.roiBtn.setStyleSheet("QPushButton { font-weight: bold; font-size: 16px;} ")
        self.configROIWindow.roiDrawBtn.clicked.connect(self.roiMerge)
        self.configROIWindow.okBtn.clicked.connect(self.closeROIWindow)
##        self.roiBtn.resize(self.roiBtn.minimumSizeHint())
##        self.roiBtn.move(375,500)

        self.measureBtn = QPushButton("Measure", self) #Measure pixel length
        self.measureBtn.clicked.connect(self.measureScale)
        self.measureBtn.setStyleSheet("QPushButton { font-weight: bold; font-size: 18px;} ")
##        self.measureBtn.resize(self.measureBtn.minimumSizeHint())
##        self.measureBtn.move(375,575)

        self.pixelValue = QDoubleSpinBox(self)
        self.pixelValue.setRange(0, 3000)
        self.pixelValue.setValue(168.67)
        self.pixelValue.setSingleStep(10)
        self.pixelValue.valueChanged.connect(self.update_calib)
##        self.pixelValue.move(375, 550)
##        self.pixelValue.resize(50, 20)

        self.pixelLabel = QLabel("px\t=\t", self) #pixel
##        self.pixelLabel.move(430, 545)
        
        self.lengthValue = QDoubleSpinBox(self)
        self.lengthValue.setRange(0, 1000)
        self.lengthValue.setValue(100)
        self.lengthValue.setSingleStep(1)
        self.lengthValue.valueChanged.connect(self.update_calib)

        self.lengthUnit = QComboBox(self) #length unit
        self.lengthUnit.addItem("px")
        self.lengthUnit.addItem("cm")
        self.lengthUnit.addItem("mm")
        self.lengthUnit.addItem("μm")
        self.lengthUnit.addItem("nm")
        self.lengthUnit.setCurrentIndex(3)
        self.lengthUnit.currentIndexChanged.connect(self.update_calib)
        
        self.videoEffect = QComboBox(self) #video effect dropdown
        self.videoEffect.addItem("Binary")
        self.videoEffect.addItem("Masked")
        self.videoEffect.addItem("Contours")
        self.videoEffect.addItem("Transformed")
        self.videoEffect.addItem("Background")
        self.videoEffect.addItem("Auto ROI")
        self.videoEffect.addItem("Force/Area Plot")
        self.videoEffect.setStyleSheet("QComboBox { font-weight: bold; font-size: 14px;} ")
##        self.videoEffect.move(650, 85)
##        self.videoEffect.resize(self.videoEffect.minimumSizeHint())
        self.videoEffect.currentIndexChanged.connect(self.effect_change)
        self.videoEffect.model().item(5).setEnabled(False)
        self.videoEffect.model().item(6).setEnabled(False)

        self.analyzeVideo = QCheckBox('Analyze Video', self) #Perform Analysis
##        self.analyzeVideo.move(500, 30)
        self.analyzeVideo.stateChanged.connect(self.anal_change)
        self.analyzeVideo.setStyleSheet("QCheckBox { font-weight: bold; font-size: 16px;} ")

        self.correctZeroForce = QCheckBox('Zero-force correct', self) #Perform Analysis
        self.correctZeroForce.stateChanged.connect(self.plotSequence)
        self.correctZeroForce.setEnabled(False)

        self.showContours = QCheckBox('Show Contours', self) #show contours
##        self.showContours.move(300, 60)
        self.showContours.stateChanged.connect(self.anal_change)
        self.showContours.setStyleSheet("QCheckBox { font-weight: bold; font-size: 16px;} ")

        self.showEffect = QCheckBox('Show Effect', self) #show contours
##        self.showContours.move(300, 60)
        self.showEffect.stateChanged.connect(self.anal_change)
        self.showEffect.setStyleSheet("QCheckBox { font-weight: bold; font-size: 16px;} ")
        
        self.bgButton = QPushButton("Set Background", self) #Set background
        self.bgButton.clicked.connect(self.bgFrame)
        self.bgButton.setStyleSheet("QPushButton { font-weight: bold; font-size: 14px;} ")
        self.bgButton.clicked.connect(self.bg_change)
##        self.bgButton.resize(self.bgButton.minimumSizeHint())
##        self.bgButton.move(420,80)

##        self.bgSubtract = QCheckBox('Subtract Background', self) #background subtract
##        self.bgSubtract.move(300, 80)
##        self.bgSubtract.stateChanged.connect(self.bg_change)

        self.backgroundCorrection =  QComboBox(self) #Background correction type
        self.backgroundCorrection.addItem("Average Correction")
        self.backgroundCorrection.addItem("Proper Correction")
        self.backgroundCorrection.addItem("Gaussian Correction")
        self.backgroundCorrection.addItem("Rolling Ball")
        self.backgroundCorrection.addItem("Rolling Paraboloid")
        self.backgroundCorrection.currentIndexChanged.connect(self.bg_change)        

        self.bgSlider = QSlider(Qt.Horizontal, self) #background blur
        self.bgSlider.setMinimum(3)
        self.bgSlider.setMaximum(1001)
        self.bgSlider.setValue(30)
        self.bgSlider.setTickInterval(200)
        self.bgSlider.setTickPosition(QSlider.TicksBelow)
        self.bgSlider.valueChanged.connect(self.bg_change)
        self.bgSlider.setSingleStep(5)
        self.bgSpinBox = QSpinBox(self)
        self.bgSpinBox.setRange(3, 1001)
        self.bgSpinBox.setValue(self.bgSlider.value())
        self.bgSpinBox.setSingleStep(1)
        self.bgSpinBox.valueChanged.connect(self.bg_change)

        self.bgAlphaLabel = QLabel("\tBlend:", self) #background alpha
        self.bgAlphaSpinBox = QDoubleSpinBox(self) 
        self.bgAlphaSpinBox.setRange(0, 1)
        self.bgAlphaSpinBox.setValue(0.45)
        self.bgAlphaSpinBox.setSingleStep(0.01)
        self.bgAlphaSpinBox.valueChanged.connect(self.bg_change)
        
        self.threshType = QComboBox(self) #threshold type
        self.threshType.addItem("Global")
        self.threshType.addItem("Adaptive")
        self.threshType.addItem("Otsu")
##        self.threshType.move(900, 35)
##        self.threshType.resize(self.threshType.minimumSizeHint())
        self.threshType.currentIndexChanged.connect(self.threshold)

        self.thresh1Label = QLabel("Cutoff:\t", self) #threshold cutoff
        self.threshSlider1 = QSlider(Qt.Horizontal, self) 
##        self.threshSlider1.setGeometry(800, 60, 200, 30)
        self.threshSlider1.setMinimum(3)
        self.threshSlider1.setMaximum(701)
        self.threshSlider1.setValue(128)
        self.threshSlider1.setTickInterval(50)
        self.threshSlider1.setTickPosition(QSlider.TicksBelow)
        self.threshSlider1.valueChanged.connect(self.threshold)
        self.threshSlider1.setSingleStep(4)
        self.threshSpinBox1 = QSpinBox(self)
        self.threshSpinBox1.setRange(3, 701)
        self.threshSpinBox1.setValue(self.threshSlider1.value())
        self.threshSpinBox1.setSingleStep(2)
        self.threshSpinBox1.valueChanged.connect(self.threshold)
##        self.threshSpinBox1.move(1010, 60)
##        self.threshSpinBox1.resize(45, 20)

        self.thresh2Label = QLabel("Constant:\t", self) #threshold constant
        self.threshSlider2 = QSlider(Qt.Horizontal, self) 
##        self.threshSlider2.setGeometry(800, 90, 200, 30)
        self.threshSlider2.setMinimum(-10)
        self.threshSlider2.setMaximum(10)
        self.threshSlider2.setValue(2)
        self.threshSlider2.setTickInterval(1)
        self.threshSlider2.setTickPosition(QSlider.TicksBelow)
        self.threshSlider2.valueChanged.connect(self.threshold)
        self.threshSlider2.setSingleStep(2)
        self.threshSpinBox2 = QSpinBox(self)
        self.threshSpinBox2.setRange(-10, 10)
        self.threshSpinBox2.setValue(self.threshSlider2.value())
        self.threshSpinBox2.setSingleStep(1)
        self.threshSpinBox2.valueChanged.connect(self.threshold)
##        self.threshSpinBox2.move(1010, 90)
##        self.threshSpinBox2.resize(45, 20)
        
        self.applySegment = QCheckBox('Segment', self) #apply image segmentation
##        self.showContours.move(300, 60)
        self.applySegment.stateChanged.connect(self.segment_change)
        
        self.showSegment = QPushButton("Display Segment", self) #display segment
        self.showSegment.setCheckable(True)
        self.showSegment.clicked.connect(self.segment_show)
        self.showSegment.setEnabled(False)

        self.segmentFGButton = QPushButton("Foreground:", self) #segment foreground parameter
        self.segmentFGButton.setCheckable(True)
        self.segmentFGButton.clicked.connect(self.segment_show_fg)
        self.segmentFGButton.setEnabled(False)
        self.segmentFGSlider = QSlider(Qt.Horizontal, self)
        self.segmentFGSlider.setMinimum(0)
        self.segmentFGSlider.setMaximum(10)
        self.segmentFGSlider.setValue(7)
        self.segmentFGSlider.setTickInterval(2)
        self.segmentFGSlider.setTickPosition(QSlider.TicksBelow)
        self.segmentFGSlider.valueChanged.connect(self.segment_param_change)
        self.segmentFGSlider.setSingleStep(2)
        self.segmentFGSlider.setEnabled(False)
        self.segmentFGSpinBox = QDoubleSpinBox(self)
        self.segmentFGSpinBox.setRange(0, 1)
        self.segmentFGSpinBox.setValue(self.segmentFGSlider.value()/10)
        self.segmentFGSpinBox.setSingleStep(0.1)
        self.segmentFGSpinBox.valueChanged.connect(self.segment_param_change)
        self.segmentFGSpinBox.setEnabled(False)

        self.segmentBGButton = QPushButton("Background:", self) #segment background parameter
        self.segmentBGButton.setCheckable(True)
        self.segmentBGButton.clicked.connect(self.segment_show_bg)
        self.segmentBGButton.setEnabled(False)
        self.segmentBGSlider = QSlider(Qt.Horizontal, self)
        self.segmentBGSlider.setMinimum(1)
        self.segmentBGSlider.setMaximum(40)
        self.segmentBGSlider.setValue(3)
        self.segmentBGSlider.setTickInterval(5)
        self.segmentBGSlider.setTickPosition(QSlider.TicksBelow)
        self.segmentBGSlider.valueChanged.connect(self.segment_param_change)
        self.segmentBGSlider.setSingleStep(2)
        self.segmentBGSlider.setEnabled(False)
        self.segmentBGSpinBox = QSpinBox(self)
        self.segmentBGSpinBox.setRange(1, 40)
        self.segmentBGSpinBox.setValue(self.segmentBGSlider.value())
        self.segmentBGSpinBox.setSingleStep(1)
        self.segmentBGSpinBox.valueChanged.connect(self.segment_param_change)
        self.segmentBGSpinBox.setEnabled(False)

        self.brightnessLabel = QLabel("Brightness:\t", self) #brightness
        self.brightnessSlider = QSlider(Qt.Horizontal, self) 
##        self.brightnessSlider.setGeometry(30, 580, 200, 30)
        self.brightnessSlider.setMinimum(-127)
        self.brightnessSlider.setMaximum(127)
        self.brightnessSlider.setValue(0)
        self.brightnessSlider.setTickInterval(50)
        self.brightnessSlider.setTickPosition(QSlider.TicksBelow)
        self.brightnessSlider.valueChanged.connect(self.bc_change)
        self.brightnessSlider.setSingleStep(5)
        self.brightnessSpinBox = QSpinBox(self)
        self.brightnessSpinBox.setRange(-127, 127)
        self.brightnessSpinBox.setValue(self.brightnessSlider.value())
        self.brightnessSpinBox.setSingleStep(1)
        self.brightnessSpinBox.valueChanged.connect(self.bc_change)
##        self.brightnessSpinBox.move(250, 580)
##        self.brightnessSpinBox.resize(45, 20)

        self.contrastLabel = QLabel("Contrast:\t", self) #contrast
        self.contrastSlider = QSlider(Qt.Horizontal, self) 
##        self.contrastSlider.setGeometry(30, 630, 200, 30)
        self.contrastSlider.setMinimum(-64)
        self.contrastSlider.setMaximum(64)
        self.contrastSlider.setValue(0)
        self.contrastSlider.setTickInterval(50)
        self.contrastSlider.setTickPosition(QSlider.TicksBelow)
        self.contrastSlider.valueChanged.connect(self.bc_change)
        self.contrastSlider.setSingleStep(5)
        self.contrastSpinBox = QSpinBox(self)
        self.contrastSpinBox.setRange(-64, 64)
        self.contrastSpinBox.setValue(self.contrastSlider.value())
        self.contrastSpinBox.setSingleStep(1)
        self.contrastSpinBox.valueChanged.connect(self.bc_change)
##        self.contrastSpinBox.move(250, 630)
##        self.contrastSpinBox.resize(45, 20)

##        self.applyFilter = QCheckBox('Apply Filter', self) #apply filter
##        self.applyFilter.move(700, 525)
##        self.applyFilter.stateChanged.connect(self.dft_change)

        self.minAreaFilter = QSpinBox(self)
        self.minAreaFilter.setRange(1, 10000)
        self.minAreaFilter.setValue(25)
        self.minAreaFilter.setSingleStep(1)
        self.minAreaFilter.valueChanged.connect(self.video_analysis)
        self.minAreaLabel = QLabel("Min Area:", self)

        self.maxAreaFilter = QSpinBox(self)
        self.maxAreaFilter.setRange(1, 1000000)
        self.maxAreaFilter.setValue(1000000)
        self.maxAreaFilter.setSingleStep(100)
        self.maxAreaFilter.valueChanged.connect(self.video_analysis)
        self.maxAreaLabel = QLabel("Max Area:", self)
        
        self.filterType =  QComboBox(self) #Filter Type
        self.filterType.addItem("Average Filter")
        self.filterType.addItem("Fourier Filter")
        self.filterType.addItem("Gaussian Filter")
        self.filterType.addItem("Median Filter")
        self.filterType.addItem("Bilateral Filter")
        self.filterType.addItem("Morph Open")
        self.filterType.addItem("Morph Close")
        self.filterType.currentIndexChanged.connect(self.dft_change)

        self.lowPassLabel = QLabel("Low Pass:\t", self) #low pass filter
        self.lowPassSlider = QSlider(Qt.Horizontal, self) 
##        self.lowPassSlider.setGeometry(800, 500, 200, 30)
        self.lowPassSlider.setMinimum(0)
        self.lowPassSlider.setMaximum(1000)
        self.lowPassSlider.setValue(6)
        self.lowPassSlider.setTickInterval(50)
        self.lowPassSlider.setTickPosition(QSlider.TicksBelow)
        self.lowPassSlider.valueChanged.connect(self.dft_change)
        self.lowPassSlider.setSingleStep(1)
        self.lowPassSpinBox = QSpinBox(self)
        self.lowPassSpinBox.setRange(0, 1000)
        self.lowPassSpinBox.setValue(self.lowPassSlider.value())
        self.lowPassSpinBox.setSingleStep(1)
        self.lowPassSpinBox.valueChanged.connect(self.dft_change)
##        self.lowPassSpinBox.move(1020, 500)
##        self.lowPassSpinBox.resize(45, 20)

        self.highPassLabel = QLabel("High Pass:\t", self) #high pass filter
        self.highPassSlider = QSlider(Qt.Horizontal, self) 
##        self.highPassSlider.setGeometry(800, 550, 200, 30)
        self.highPassSlider.setMinimum(0)
        self.highPassSlider.setMaximum(1000)
        self.highPassSlider.setValue(0)
        self.highPassSlider.setTickInterval(50)
        self.highPassSlider.setTickPosition(QSlider.TicksBelow)
        self.highPassSlider.valueChanged.connect(self.dft_change)
        self.highPassSlider.setSingleStep(5)
        self.highPassSpinBox = QSpinBox(self)
        self.highPassSpinBox.setRange(0, 1000)
        self.highPassSpinBox.setValue(self.highPassSlider.value())
        self.highPassSpinBox.setSingleStep(1)
        self.highPassSpinBox.valueChanged.connect(self.dft_change)
##        self.highPassSpinBox.move(1020, 550)
##        self.highPassSpinBox.resize(45, 20)

##        self.autoDetectROI = QCheckBox('Auto detect ROI', self) #Auto detect ROI
##        self.autoDetectROI.move(430, 605)
##        self.autoDetectROI.stateChanged.connect(self.auto_roi_change)

        self.distinctAutoROI = QCheckBox('Distinct ROI', self) #Distinct auto ROI (one for each manual roi)
        self.distinctAutoROI.setChecked(True)
        self.distinctAutoROI.stateChanged.connect(self.distinct_roi_change)

        self.combineROI = QCheckBox('Combine', self) #combine multiple contour data sets of ROI
        self.combineROI.stateChanged.connect(self.distinct_roi_change)

        self.morphROI = QCheckBox('Morph', self) #Morph ROI (remove holes)
        self.morphROI.setChecked(True)
        self.morphROI.stateChanged.connect(self.roi_morph_change)

        self.morphXLabel = QLabel("Morph X:\t", self)
        self.morphXSpinBox = QSpinBox(self) #morph kernel X
        self.morphXSpinBox.setRange(1, 1000)
        self.morphXSpinBox.setValue(28)
        self.morphXSpinBox.setSingleStep(1)
        self.morphXSpinBox.valueChanged.connect(self.roi_morph_change)
        # self.morphXSpinBox.setEnabled(False)

        self.morphYLabel = QLabel("Morph Y:\t", self)
        self.morphYSpinBox = QSpinBox(self) #morph kernel Y
        self.morphYSpinBox.setRange(1, 1000)
        self.morphYSpinBox.setValue(62)
        self.morphYSpinBox.setSingleStep(1)
        self.morphYSpinBox.valueChanged.connect(self.roi_morph_change)
        # self.morphYSpinBox.setEnabled(False)
        
        self.applyHullROI = QCheckBox('Apply Hull', self) #Apply convex hull on ROI
        self.applyHullROI.stateChanged.connect(self.epsilon_change)

        self.applybgROI = QCheckBox('BG Correct', self) #Apply background correction on ROI
        self.applybgROI.stateChanged.connect(self.bg_roi_change)
        
        self.threshROIType = QComboBox(self) #auto roi threshold type
        self.threshROIType.addItem("Global")
        self.threshROIType.addItem("Adaptive")
        self.threshROIType.addItem("Otsu")
        self.threshROIType.addItem("Canny")
##        self.threshROIType.move(550, 610)
##        self.threshROIType.resize(self.threshROIType.minimumSizeHint())
        self.threshROIType.currentIndexChanged.connect(self.threshold_roi)
        
        self.epsilonROILabel = QLabel("Resolution:\t", self) #adjust resolution of roi polyfit
        self.epsilonSlider = QSlider(Qt.Horizontal, self) 
##        self.epsilonSlider.setGeometry(400, 650, 200, 30)
        self.epsilonSlider.setMinimum(-50)
        self.epsilonSlider.setMaximum(0)
        self.epsilonSlider.setValue(-30)
        self.epsilonSlider.setTickInterval(10)
        self.epsilonSlider.setTickPosition(QSlider.TicksBelow)
        self.epsilonSlider.valueChanged.connect(self.epsilon_change)
        self.epsilonSlider.setSingleStep(5)
        self.epsilonSpinBox = QSpinBox(self)
        self.epsilonSpinBox.setRange(-50, 0)
        self.epsilonSpinBox.setValue(self.epsilonSlider.value())
        self.epsilonSpinBox.setSingleStep(1)
        self.epsilonSpinBox.valueChanged.connect(self.epsilon_change)
##        self.epsilonSpinBox.move(620, 650)
##        self.epsilonSpinBox.resize(45, 20)

        self.roiMinLabel = QLabel("Minimum:", self) #minimum area of roi contours
        self.roiMinSpinBox = QSpinBox(self)
        self.roiMinSpinBox.setRange(0, 1000)
        self.roiMinSpinBox.setValue(3)
        self.roiMinSpinBox.setSingleStep(1)
        self.roiMinSpinBox.valueChanged.connect(self.roi_min_change)

        self.resizeROILabel = QLabel("Resize factor:", self) #resize roi
        self.resizeROISpinBox = QDoubleSpinBox(self)
        self.resizeROISpinBox.setRange(0, 10)
        self.resizeROISpinBox.setValue(0.96)
        self.resizeROISpinBox.setSingleStep(0.1)
        self.resizeROISpinBox.valueChanged.connect(self.video_analysis)
        
        self.threshROI1Label = QLabel("Cutoff:\t", self) #threshold cutoff (roi)
        self.threshROISlider1 = QSlider(Qt.Horizontal, self)
##        self.threshROISlider1.setGeometry(400, 700, 200, 30)
        self.threshROISlider1.setMinimum(1)
        self.threshROISlider1.setMaximum(1001)
        self.threshROISlider1.setValue(148)
        self.threshROISlider1.setTickInterval(50)
        self.threshROISlider1.setTickPosition(QSlider.TicksBelow)
        self.threshROISlider1.valueChanged.connect(self.threshold_roi)
        self.threshROISlider1.setSingleStep(4)
        self.threshROISpinBox1 = QSpinBox(self)
        self.threshROISpinBox1.setRange(1, 1001)
        self.threshROISpinBox1.setValue(self.threshROISlider1.value())
        self.threshROISpinBox1.setSingleStep(2)
        self.threshROISpinBox1.valueChanged.connect(self.threshold_roi)
##        self.threshROISpinBox1.move(620, 700)
##        self.threshROISpinBox1.resize(45, 20)

        self.threshROI2Label = QLabel("Constant:\t", self) #threshold constant (roi)
        self.threshROISlider2 = QSlider(Qt.Horizontal, self)
##        self.threshROISlider2.setGeometry(400, 750, 200, 30)
        self.threshROISlider2.setMinimum(-10)
        self.threshROISlider2.setMaximum(700)
        self.threshROISlider2.setValue(2)
        self.threshROISlider2.setTickInterval(50)
        self.threshROISlider2.setTickPosition(QSlider.TicksBelow)
        self.threshROISlider2.valueChanged.connect(self.threshold_roi)
        self.threshROISlider2.setSingleStep(2)
        self.threshROISpinBox2 = QSpinBox(self)
        self.threshROISpinBox2.setRange(-10, 700)
        self.threshROISpinBox2.setValue(self.threshROISlider2.value())
        self.threshROISpinBox2.setSingleStep(1)
        self.threshROISpinBox2.valueChanged.connect(self.threshold_roi)
##        self.threshROISpinBox2.move(620, 750)
##        self.threshROISpinBox2.resize(45, 20)

        self.blurROILabel = QLabel("Blur:\t", self) #blur for auto roi detection (roi)
        self.blurROISlider = QSlider(Qt.Horizontal, self)
##        self.threshROISlider2.setGeometry(400, 750, 200, 30)
        self.blurROISlider.setMinimum(3)
        self.blurROISlider.setMaximum(1000)
        self.blurROISlider.setValue(7)
        self.blurROISlider.setTickInterval(100)
        self.blurROISlider.setTickPosition(QSlider.TicksBelow)
        self.blurROISlider.valueChanged.connect(self.blur_roi)
        self.blurROISlider.setSingleStep(2)
        self.blurROISpinBox = QSpinBox(self)
        self.blurROISpinBox.setRange(3, 1000)
        self.blurROISpinBox.setValue(self.blurROISlider.value())
        self.blurROISpinBox.setSingleStep(1)
        self.blurROISpinBox.valueChanged.connect(self.blur_roi)
##        self.threshROISpinBox2.move(620, 750)
##        self.threshROISpinBox2.resize(45, 20)

        self.bgblurROILabel = QLabel("BG Blur:\t", self) #blur background for auto roi detection (roi)
        self.bgblurROISlider = QSlider(Qt.Horizontal, self)
##        self.threshROISlider2.setGeometry(400, 750, 200, 30)
        self.bgblurROISlider.setMinimum(3)
        self.bgblurROISlider.setMaximum(1000)
        self.bgblurROISlider.setValue(3)
        self.bgblurROISlider.setTickInterval(100)
        self.bgblurROISlider.setTickPosition(QSlider.TicksBelow)
        self.bgblurROISlider.valueChanged.connect(self.bg_blur_roi)
        self.bgblurROISlider.setSingleStep(2)
        self.bgblurROISlider.setEnabled(False)
        self.bgblurROISpinBox = QSpinBox(self)
        self.bgblurROISpinBox.setRange(3, 1000)
        self.bgblurROISpinBox.setValue(self.bgblurROISlider.value())
        self.bgblurROISpinBox.setSingleStep(1)
        self.bgblurROISpinBox.valueChanged.connect(self.bg_blur_roi)
        self.bgblurROISpinBox.setEnabled(False)

        self.bgblendROILabel = QLabel("BG Blend:\t", self) #blend background for auto roi detection (roi)
        self.bgblendROISlider = QSlider(Qt.Horizontal, self)
##        self.threshROISlider2.setGeometry(400, 750, 200, 30)
        self.bgblendROISlider.setMinimum(0)
        self.bgblendROISlider.setMaximum(100)
        self.bgblendROISlider.setValue(50)
        self.bgblendROISlider.setTickInterval(10)
        self.bgblendROISlider.setTickPosition(QSlider.TicksBelow)
        self.bgblendROISlider.valueChanged.connect(self.bg_blend_roi)
        self.bgblendROISlider.setSingleStep(10)
        self.bgblendROISlider.setEnabled(False)
        self.bgblendROISpinBox = QDoubleSpinBox(self)
        self.bgblendROISpinBox.setRange(0, 1)
        self.bgblendROISpinBox.setValue(self.bgblendROISlider.value()/100)
        self.bgblendROISpinBox.setSingleStep(0.01)
        self.bgblendROISpinBox.valueChanged.connect(self.bg_blend_roi)
        self.bgblendROISpinBox.setEnabled(False)
        
        self.effectContrast = QCheckBox('Enhance Contrast', self) #enhance effect contrast
        self.effectContrast.stateChanged.connect(self.video_analysis)

        self.showPlot = QPushButton("Live Plot", self) #Show Plot
        self.showPlot.clicked.connect(self.plot_data)
##        self.showPlot.setSizePolicy(QSizePolicy.Minimum, QSizePolicy.Minimum)
        self.showPlot.setStyleSheet("QPushButton { font-weight: bold; font-size: 20px;} ")
##        self.showPlot.resize(self.showPlot.minimumSizeHint())
##        self.showPlot.move(800,600)

        self.saveBtn = QPushButton("Save Data", self) #Save area data
        self.saveBtn.clicked.connect(self.save_data)
        # self.saveBtn.setSizePolicy(QSizePolicy.Minimum, QSizePolicy.Minimum)
        self.saveBtn.setStyleSheet("QPushButton { font-weight: bold; font-size: 24px;} ")
##        self.saveBtn.resize(self.saveBtn.minimumSizeHint())
##        self.saveBtn.move(900,600)

        self.clearData = QPushButton("Clear Data", self) #Clear area data
        self.clearData.clicked.connect(self.clear_data)
        # self.clearData.setSizePolicy(QSizePolicy.Minimum, QSizePolicy.Minimum)
        self.clearData.setStyleSheet("QPushButton { font-weight: bold; font-size: 24px;} ")
##        self.clearData.resize(self.clearData.minimumSizeHint())
##        self.clearData.move(900,650)
        self.clearData.setEnabled(False)

        self.fpsSpinBox = QDoubleSpinBox(self) #fps input
        self.fpsSpinBox.setMinimum(0)
        self.fpsSpinBox.setValue(1)
##        self.fpsSpinBox.move(650,55)
##        self.fpsSpinBox.resize(45, 20)
        self.fpsSpinBox.valueChanged.connect(self.fps_change)
        self.fpsLabel = QLabel("Enter FPS", self)
##        self.fpsLabel.move(595,50)
        
        self.videoFileNameLabel = QLabel("Select video from file menu", self)
##        self.videoFileNameLabel.setGeometry(10, 650, 400, 60)
##        self.videoFileNameLabel.setTextFormat(Qt.RichText)
        self.videoFileNameLabel.setWordWrap(True)

        self.forceFileNameLabel = QLabel("Select force data from file menu", self)
##        self.forceFileNameLabel.setGeometry(10, 700, 400, 60)
        self.forceFileNameLabel.setWordWrap(True)

        self.zeroForceFileNameLabel = QLabel("Select zero-force data from file menu", self)
        self.zeroForceFileNameLabel.setWordWrap(True)
        

        ################ Adjust window layout #######################
        
        wid = QWidget(self)
        self.setCentralWidget(wid)

        topleftGroupBox = QGroupBox()
        # topleftGroupBox.setSizePolicy(QSizePolicy.Ignored, QSizePolicy.Ignored)
        topleftGroupBox.setStyleSheet("QGroupBox { font-weight: bold; } ")
        topleftVbox = QGridLayout(self)
        topleftVbox.setColumnStretch(0, 1.5)
        topleftGroupBox.setLayout(topleftVbox)
        topleftVbox.addWidget(self.showContours, 0, 0, 1, 1)
        topleftVbox.addWidget(self.showEffect, 0, 1, 1, 1)
        topleftVbox.addWidget(self.roiBtn, 0, 2, 1, 1)

        toprightGroupBox = QGroupBox()
        # toprightGroupBox.setSizePolicy(QSizePolicy.Ignored, QSizePolicy.Ignored)
        toprightGroupBox.setStyleSheet("QGroupBox { font-weight: bold; } ")
        toprightVbox = QGridLayout(self)
        toprightVbox.setColumnStretch(0, 1.5)
        toprightGroupBox.setLayout(toprightVbox)
        toprightVbox.addWidget(self.analyzeVideo, 0, 0, 1, 1)
        toprightVbox.addWidget(self.videoEffect, 0, 1, 1, 1)

        displayGroupBox = QGroupBox()
        # displayGroupBox.setSizePolicy(QSizePolicy.Minimum, QSizePolicy.Minimum)
        displayGroupBox.setStyleSheet("QGroupBox { font-weight: bold; } ")
        displayVbox = QGridLayout(self)
        displayGroupBox.setLayout(displayVbox)
        displayVbox.addWidget(self.rawView, 0, 0, 1, 1)
        displayVbox.addWidget(self.effectView, 0, 1, 1, 1)

        playbackGroupBox = QGroupBox()
        # playbackGroupBox.setSizePolicy(QSizePolicy.Ignored, QSizePolicy.Ignored)
        playbackGroupBox.setStyleSheet("QGroupBox { font-weight: bold; } ")
        playbackVbox = QGridLayout(self)
        playbackVbox.setRowStretch(1, 1.5)
        playbackGroupBox.setLayout(playbackVbox)
        playbackVbox.addWidget(self.seekSlider, 0, 0, 1, 5)
        playbackVbox.addWidget(self.playBtn, 1, 0, 1, 1)
        playbackVbox.addWidget(self.stopBtn, 1, 1, 1, 1)
        playbackVbox.addWidget(self.prevBtn, 1, 2, 1, 1)
        playbackVbox.addWidget(self.nextBtn, 1, 3, 1, 1)
        playbackVbox.addWidget(self.recordBtn, 1, 4, 1, 1)        

        self.bgGroupBox = QGroupBox("Background Correction")
        # self.bgGroupBox.setSizePolicy(QSizePolicy.Ignored, QSizePolicy.Ignored)
        self.bgGroupBox.setStyleSheet("QGroupBox { font-weight: bold; } ")
        self.bgGroupBox.setCheckable(True)
        self.bgGroupBox.setChecked(False)        
        self.bgGroupBox.toggled.connect(self.bg_change)
        bgVbox = QGridLayout(self)
        self.bgGroupBox.setLayout(bgVbox)
##        bgVbox.addWidget(self.bgSubtract, 0, 0, 1, 1)
        bgVbox.addWidget(self.bgButton, 1, 0, 1, 1)
        bgVbox.addWidget(self.backgroundCorrection, 0, 0, 1, 1)
        bgVbox.addWidget(self.bgSlider, 0, 1, 1, 1)
        bgVbox.addWidget(self.bgSpinBox, 0, 2, 1, 1)
        bgVbox.addWidget(self.bgAlphaLabel, 1, 1, 1, 1)
        bgVbox.addWidget(self.bgAlphaSpinBox, 1, 2, 1, 1)
        
        
        measureGroupBox = QGroupBox("Length Calibration")
        # measureGroupBox.setSizePolicy(QSizePolicy.Ignored, QSizePolicy.Ignored)
        measureGroupBox.setStyleSheet("QGroupBox { font-weight: bold; } ")
        measureVbox = QGridLayout(self)
        measureGroupBox.setLayout(measureVbox)
        measureVbox.addWidget(self.pixelValue, 0, 0, 1, 1)
        measureVbox.addWidget(self.pixelLabel, 0, 1, 1, 1)
        measureVbox.addWidget(self.lengthValue, 0, 2, 1, 1)
        measureVbox.addWidget(self.lengthUnit, 0, 3, 1, 1)
        measureVbox.addWidget(self.measureBtn, 1, 1, 1, 2)

        fpsGroupBox = QGroupBox("Adjust")
        # fpsGroupBox.setSizePolicy(QSizePolicy.Ignored, QSizePolicy.Ignored)
        fpsGroupBox.setStyleSheet("QGroupBox { font-weight: bold; } ")
        fpsVbox = QGridLayout(self)
        fpsGroupBox.setLayout(fpsVbox)
        fpsVbox.addWidget(self.fpsLabel, 0, 0, 1, 1)
        fpsVbox.addWidget(self.fpsSpinBox, 0, 1, 1, 1)
        fpsVbox.addWidget(self.correctZeroForce, 1, 0, 1, 2)

        self.middleleftGroupBox = QGroupBox()
        # self.middleleftGroupBox.setSizePolicy(QSizePolicy.Ignored, QSizePolicy.Ignored)
        middleleftVbox = QGridLayout(self)
        self.middleleftGroupBox.setLayout(middleleftVbox)
        middleleftVbox.addWidget(self.bgGroupBox, 0, 0, 1, 2)
        middleleftVbox.addWidget(measureGroupBox, 0, 2, 1, 2)
        middleleftVbox.addWidget(fpsGroupBox, 0, 4, 1, 1)

        self.bcGroupBox = QGroupBox("Brightness/Contrast")
        # self.bcGroupBox.setSizePolicy(QSizePolicy.Ignored, QSizePolicy.Ignored)
        self.bcGroupBox.setStyleSheet("QGroupBox { font-weight: bold; } ")
        bcVbox = QGridLayout(self)
        self.bcGroupBox.setLayout(bcVbox)
        bcVbox.addWidget(self.brightnessLabel, 0, 0, 1, 1)
        bcVbox.addWidget(self.brightnessSlider, 0, 1, 1, 1)
        bcVbox.addWidget(self.brightnessSpinBox, 0, 3, 1, 1)
        bcVbox.addWidget(self.contrastLabel, 1, 0, 1, 1)
        bcVbox.addWidget(self.contrastSlider, 1, 1, 1, 1)
        bcVbox.addWidget(self.contrastSpinBox, 1, 3, 1, 1)

        self.dftGroupBox = QGroupBox("Image Filtering")
        # self.dftGroupBox.setSizePolicy(QSizePolicy.Ignored, QSizePolicy.Ignored)
        self.dftGroupBox.setStyleSheet("QGroupBox { font-weight: bold; } ")
        self.dftGroupBox.setCheckable(True)
        self.dftGroupBox.setChecked(False)        
        self.dftGroupBox.toggled.connect(self.dft_change)
        dftVbox = QGridLayout(self)
        self.dftGroupBox.setLayout(dftVbox)
##        dftVbox.addWidget(self.applyFilter, 0, 0, 1, 1)
        dftVbox.addWidget(self.filterType, 0, 0, 1, 1)
        dftVbox.addWidget(self.maxAreaLabel, 0, 7, 1, 1)
        dftVbox.addWidget(self.maxAreaFilter, 0, 8, 1, 1)
        dftVbox.addWidget(self.minAreaLabel, 0, 9, 1, 1)
        dftVbox.addWidget(self.minAreaFilter, 0, 10, 1, 1)
        dftVbox.addWidget(self.lowPassLabel, 1, 0, 1, 1)
        dftVbox.addWidget(self.lowPassSlider, 1, 1, 1, 9)
        dftVbox.addWidget(self.lowPassSpinBox, 1, 10, 1, 1)
        dftVbox.addWidget(self.highPassLabel, 2, 0, 1, 1)
        dftVbox.addWidget(self.highPassSlider, 2, 1, 1, 9)
        dftVbox.addWidget(self.highPassSpinBox, 2, 10, 1, 1)

        fileGroupBox = QGroupBox("Filenames")
        # fileGroupBox.setSizePolicy(QSizePolicy.Ignored, QSizePolicy.Ignored)
        fileGroupBox.setStyleSheet("QGroupBox { font-weight: bold; } ")
        fileVbox = QGridLayout(self)
        fileGroupBox.setLayout(fileVbox)
        fileVbox.addWidget(self.videoFileNameLabel, 0, 0, 1, 1)
        fileVbox.addWidget(self.forceFileNameLabel, 1, 0, 1, 1)
        fileVbox.addWidget(self.zeroForceFileNameLabel, 2, 0, 1, 1)

        middlerightGroupBox = QGroupBox()
        # middlerightGroupBox.setSizePolicy(QSizePolicy.Ignored, QSizePolicy.Ignored)
        middlerightVbox = QGridLayout(self)
        middlerightVbox.setColumnStretch(0, 1)
        middlerightVbox.setColumnStretch(2, 1.5)
        middlerightGroupBox.setLayout(middlerightVbox)
        middlerightVbox.addWidget(self.showPlot, 0, 1, 1, 1)
        middlerightVbox.addWidget(self.effectContrast, 0, 0, 1, 1)
        
        self.threshGroupBox = QGroupBox("Threshold")
        # self.threshGroupBox.setSizePolicy(QSizePolicy.Ignored, QSizePolicy.Ignored)
        self.threshGroupBox.setStyleSheet("QGroupBox { font-weight: bold; } ")
        threshVbox = QGridLayout(self)
        self.threshGroupBox.setLayout(threshVbox)
        threshVbox.addWidget(self.applySegment, 0, 0, 1, 1)
        threshVbox.addWidget(self.showSegment, 0, 1, 1, 1)
        threshVbox.addWidget(self.threshType, 0, 9, 1, 1)
        threshVbox.addWidget(self.thresh1Label, 1, 0, 1, 1)
        threshVbox.addWidget(self.threshSlider1, 1, 1, 1, 8)
        threshVbox.addWidget(self.threshSpinBox1, 1, 9, 1, 1)
        threshVbox.addWidget(self.thresh2Label, 2, 0, 1, 1)
        threshVbox.addWidget(self.threshSlider2, 2, 1, 1, 8)
        threshVbox.addWidget(self.threshSpinBox2, 2, 9, 1, 1)
        threshVbox.addWidget(self.segmentFGButton, 3, 0, 1, 1)
        threshVbox.addWidget(self.segmentFGSlider, 3, 1, 1, 3)
        threshVbox.addWidget(self.segmentFGSpinBox, 3, 4, 1, 1)
        threshVbox.addWidget(self.segmentBGButton, 3, 5, 1, 1)
        threshVbox.addWidget(self.segmentBGSlider, 3, 6, 1, 3)
        threshVbox.addWidget(self.segmentBGSpinBox, 3, 9, 1, 1)

        self.threshROIGroupBox = QGroupBox("Auto ROI Detect")
        # self.threshROIGroupBox.setSizePolicy(QSizePolicy.Ignored, QSizePolicy.Ignored)
        self.threshROIGroupBox.setStyleSheet("QGroupBox { font-weight: bold; } ")
        self.threshROIGroupBox.setCheckable(True)
        self.threshROIGroupBox.setChecked(False)        
        self.threshROIGroupBox.toggled.connect(self.auto_roi_change)
        threshROIVbox = QGridLayout(self)
        self.threshROIGroupBox.setLayout(threshROIVbox)
##        threshROIVbox.addWidget(self.autoDetectROI, 0, 0, 1, 1)
        threshROIVbox.addWidget(self.distinctAutoROI, 0, 0, 1, 1)
        threshROIVbox.addWidget(self.combineROI, 0, 1, 1, 1)
        threshROIVbox.addWidget(self.applyHullROI, 0, 2, 1, 1)
        threshROIVbox.addWidget(self.applybgROI, 0, 3, 1, 1)
        threshROIVbox.addWidget(self.roiMinLabel, 0, 5, 1, 1)
        threshROIVbox.addWidget(self.roiMinSpinBox, 0, 6, 1, 1)
        threshROIVbox.addWidget(self.resizeROILabel, 0, 7, 1, 1)
        threshROIVbox.addWidget(self.resizeROISpinBox, 0, 8, 1, 1)        
        threshROIVbox.addWidget(self.threshROIType, 0, 9, 1, 1)
        threshROIVbox.addWidget(self.epsilonROILabel, 1, 0, 1, 1)
        threshROIVbox.addWidget(self.epsilonSlider, 1, 1, 1, 3)
        threshROIVbox.addWidget(self.epsilonSpinBox, 1, 4, 1, 1)
        threshROIVbox.addWidget(self.morphROI, 1, 5, 1, 1)
        threshROIVbox.addWidget(self.morphXLabel, 1, 6, 1, 1)
        threshROIVbox.addWidget(self.morphXSpinBox, 1, 7, 1, 1)
        threshROIVbox.addWidget(self.morphYLabel, 1, 8, 1, 1)
        threshROIVbox.addWidget(self.morphYSpinBox, 1, 9, 1, 1)
        threshROIVbox.addWidget(self.threshROI1Label, 2, 0, 1, 1)
        threshROIVbox.addWidget(self.threshROISlider1, 2, 1, 1, 8)
        threshROIVbox.addWidget(self.threshROISpinBox1, 2, 9, 1, 1)
        threshROIVbox.addWidget(self.threshROI2Label, 3, 0, 1, 1)
        threshROIVbox.addWidget(self.threshROISlider2, 3, 1, 1, 8)
        threshROIVbox.addWidget(self.threshROISpinBox2, 3, 9, 1, 1)
        threshROIVbox.addWidget(self.blurROILabel, 4, 0, 1, 1)
        threshROIVbox.addWidget(self.blurROISlider, 4, 1, 1, 8)
        threshROIVbox.addWidget(self.blurROISpinBox, 4, 9, 1, 1)
        threshROIVbox.addWidget(self.bgblurROILabel, 5, 0, 1, 1)
        threshROIVbox.addWidget(self.bgblurROISlider, 5, 1, 1, 3)
        threshROIVbox.addWidget(self.bgblurROISpinBox, 5, 4, 1, 1)
        threshROIVbox.addWidget(self.bgblendROILabel, 5, 5, 1, 1)
        threshROIVbox.addWidget(self.bgblendROISlider, 5, 6, 1, 3)
        threshROIVbox.addWidget(self.bgblendROISpinBox, 5, 9, 1, 1)
        
        self.dataGroupBox = QGroupBox()
        # self.dataGroupBox.setSizePolicy(QSizePolicy.Ignored, QSizePolicy.Ignored)
        self.dataGroupBox.setStyleSheet("QGroupBox { font-weight: bold; } ")
        dataVbox = QGridLayout(self)
        self.dataGroupBox.setLayout(dataVbox)
        dataVbox.addWidget(self.saveBtn, 0, 0, 1, 1)
        dataVbox.addWidget(self.clearData, 0, 1, 1, 1)


        self.layout.addWidget(topleftGroupBox, 0, 0, 1, 1)
        self.layout.addWidget(toprightGroupBox, 0, 1, 1, 1)
        
        self.layout.addWidget(self.rawView, 1, 0, 1, 1)
        self.layout.addWidget(self.effectView, 1, 1, 1, 1)
        
        self.layout.addWidget(playbackGroupBox, 5, 0, 1, 1)
        self.layout.addWidget(middlerightGroupBox, 5, 1, 1, 1)
        
        self.layout.addWidget(self.middleleftGroupBox, 6, 0, 1, 1)
        self.layout.addWidget(self.bcGroupBox, 8, 0, 2, 1)
        self.layout.addWidget(self.dftGroupBox, 10, 0, 3, 1)

        self.layout.addWidget(fileGroupBox, 13, 0, 4, 1)

        self.layout.addWidget(self.threshGroupBox, 6, 1, 3, 1)
        self.layout.addWidget(self.threshROIGroupBox, 9, 1, 5, 1)

        self.layout.addWidget(self.dataGroupBox, 14, 1, 3, 1)

        wid.setLayout(self.layout)

        ###################### layout design finish#########################
        
        self.show()

        #scale fit graphic view
        self.rawView.fitInView(self.rawPixmapItem, 0)
        self.effectView.fitInView(self.effectPixmapItem, 0)

        self.frame = []
        self.frameAction = "Stop"
        self.videoPath = ""
        self.msrListMode = False
        self.msrmnt_num = None
        self.comb = False

        self.definePaths()

    def load_image(self): #load image file
        self.videoPath, _ = QFileDialog.getOpenFileName(self, "Open Image File")
        if self.videoPath != "":
            self.frame = cv2.imread(self.videoPath)
            self.ret = True
            self.playStatus = False
            self.recordStatus = False
            self.frameHeight, self.frameWidth = self.frame.shape[:2]
            roiCorners = np.array([[0, 0],[self.frameWidth, 0], 
                                        [self.frameWidth, self.frameHeight], 
                                        [0, self.frameHeight]],np.int32)
            self.roiBound = [0, 0, self.frameWidth, self.frameHeight]
            self.roiDict = {"Default": [roiCorners, self.roiBound, [], roiCorners, roiCorners]}
            self.frameBackground = 255 * np.ones((self.frameHeight,
                                                self.frameWidth, 3),dtype=np.uint8)
            self.renderVideo("Raw", self.ret, self.frame)
            self.frame_current = self.frame.copy()
            #Frame no. ROI label, contour id, area, length, ecc, array
            self.contour_data = [[], [], [], [], [], [], [], []] 
            contactArea = np.zeros(1, np.float64)
            contactLength = np.zeros(1, np.float64)
            contourNumber = np.zeros(1, np.uint64)
            roiArea = np.zeros(1, np.float64)
            roiLength = np.zeros(1, np.float64)
            self.frameTime = np.zeros(1, np.float64)
            eccAvg = np.zeros(1, np.float64)
            self.dataDict = {"Default" : [contactArea, contactLength,
                                        contourNumber, roiArea,
                                        roiLength, eccAvg, [(0,0),(0,0),0,1]]}
            self.effectChain = [1, 1, 1, 1] #b/c, bg sub, filter, tresh
##            self.roi_auto = self.threshROIGroupBox.isChecked()
            self.distinct_roi = self.distinctAutoROI.isChecked()
            self.roi_hull = self.applyHullROI.isChecked()
            self.combine_roi = self.combineROI.isChecked()
            self.roi_tresh_type = "Global"
            self.tresh_size_roi = self.threshROISpinBox1.value()
            self.tresh_cst_roi = self.threshROISpinBox2.value()
            self.blur_size_roi = self.blurROISpinBox.value()
            self.bg_roi_apply = self.applybgROI.isChecked()
            self.bg_blur_size_roi = self.bgblurROISpinBox.value()
            self.bg_blend_roi = self.bgblendROISpinBox.value()
            self.framePos = 0
            self.frameCount = 1
            self.roi_min_area = self.roiMinSpinBox.value()
            self.epsilon_fraction = 10**(self.epsilonSpinBox.value())
            self.roi_morph = self.morphROI.isChecked()
            self.x_roi_morph = self.morphXSpinBox.value()
            self.y_roi_morph = self.morphYSpinBox.value()
            self.calibFactor = self.lengthValue.value()/self.pixelValue.value()
            
            self.definePaths()
            
            self.analyzeVideo.blockSignals(True)
            self.analyzeVideo.setChecked(False)
            self.analyzeVideo.blockSignals(False)
            
            self.dftGroupBox.blockSignals(True)
            self.dftGroupBox.setChecked(False)
            self.dftGroupBox.blockSignals(False)

            self.bgGroupBox.blockSignals(True)
            self.bgGroupBox.setChecked(False)
            self.bgGroupBox.blockSignals(False)

            self.showContours.blockSignals(True)
            self.showContours.setChecked(False)
            self.showContours.blockSignals(False)

            self.threshROIGroupBox.blockSignals(True)
            self.threshROIGroupBox.setChecked(False)
            self.threshROIGroupBox.blockSignals(False)

##            self.distinctAutoROI.setEnabled(False)
            self.roi_auto = False
            self.segment = False
            self.show_segment = self.showSegment.isChecked()
            self.show_fg = self.segmentFGButton.isChecked()
            self.show_bg = self.segmentBGButton.isChecked()
##            self.distinct_roi = False
            
##            self.distinctAutoROI.blockSignals(True)
##            self.distinctAutoROI.setChecked(False)
##            self.distinctAutoROI.blockSignals(False)

            self.correctZeroForce.blockSignals(True)
            self.correctZeroForce.setChecked(False)
            self.correctZeroForce.blockSignals(False)


            
            self.init_dict() #initialise dictionaries
                        
            self.forceData = ForceAnal()
            self.zeroForceData = ForceAnal() #initalise
            self.correctZeroForce.setEnabled(False)
            self.clearData.setEnabled(True)
            self.videoEffect.setCurrentIndex(0)
            self.videoEffect.model().item(5).setEnabled(False)
            self.videoEffect.model().item(6).setEnabled(True)
            
            self.videoFileNameLabel.setText("<b>Video file:</b>\n" + self.videoPath) #video path
            self.forceFileNameLabel.setText("Select force data from file menu")
            self.zeroForceFileNameLabel.setText("Select zero-force data from file menu")
            self.statusBar.showMessage("Frame number: " + str(int(self.framePos)))
            
            self.frame_bin_roi_full = None
            self.setWindowTitle(self.appVersion)
            self.msrmnt_num = None
           
    def load_video(self): #load video
        if self.msrListMode == False:
            self.videoPath, _ = QFileDialog.getOpenFileName(self, "Open Video File")
        if self.videoPath != "":
##            self.statusBar.showMessage("Loading...")
            self.cap = cv2.VideoCapture(self.videoPath)
            self.playStatus = False
            self.frameAction = "Play"
##            self.frameCount = self.cap.get(cv2.CAP_PROP_FRAME_COUNT)

            self.countThread = CountFrameThread(self.cap) #count frames in separate thread
            self.countThread.output.connect(self.loading_indicate)
            self.countThread.finished.connect(self.setVideoParam)
            self.countThread.start()

    def setVideoParam(self): #set video parameters
##            self.count_frames() #count number of frames by loop
            self.cap.set(cv2.CAP_PROP_POS_FRAMES, 0)
            self.frameCount = self.countThread.frameCount
            self.frameWidth = int(self.cap.get(cv2.CAP_PROP_FRAME_WIDTH))
            self.frameHeight = int(self.cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
            self.frameRate = self.fpsSpinBox.value()
             
            print(self.frameCount,self.frameWidth, self.frameRate)
            
            roiCorners = np.array([[0, 0],[self.frameWidth, 0], 
                                        [self.frameWidth, self.frameHeight], 
                                        [0, self.frameHeight]],np.int32)
            self.roiBound = [0, 0, self.frameWidth, self.frameHeight]
            #roi dictionary: actual roi, bound, contours, adjusted roi, auto roi
            self.roiDict = {"Default": [roiCorners, self.roiBound,
                                        [], roiCorners, roiCorners]} 
            self.frameBackground = 255 * np.ones((self.frameHeight,
                                                self.frameWidth, 3),dtype=np.uint8)
            self.bgframeNumber = None
            self.ret, self.frame = self.cap.read()
            self.renderVideo("Raw", self.ret, self.frame)

            self.effectScene.removeItem(self.effectPixmapItem)
            self.effectPixmapItem = self.effectScene.addPixmap(self.blankPixmap)
            self.effectView.fitInView(self.rawPixmapItem, 1)
                
            self.cap.set(cv2.CAP_PROP_POS_FRAMES, 0)
            self.frame_current = self.frame.copy()
            #Frame no. ROI label, contour id, area, length, ecc, array
            self.contour_data = [[], [], [], [], [], [], [], []]
            contactArea = np.zeros(int(self.frameCount), np.float64)
            contactLength = np.zeros(int(self.frameCount), np.float64)
            contourNumber = np.zeros(int(self.frameCount), np.uint64)
            roiArea = np.zeros(int(self.frameCount), np.float64)
            roiLength = np.zeros(int(self.frameCount), np.float64)
            eccAvg = np.zeros(int(self.frameCount), np.float64) #median eccentricity
            self.dataDict = {"Default" : [contactArea, contactLength,
                                        contourNumber, roiArea,
                                        roiLength, eccAvg] + [[[(0,0),(0,0),0,1]]*int(self.frameCount)]}
            self.frameTime = np.linspace(0,
                                         self.frameCount/self.frameRate,
                                         int(self.frameCount), dtype = np.float64)
            self.effectChain = [1, 1, 1, 1] #b/c, bg sub, filter, tresh
##            self.roi_auto = self.threshROIGroupBox.isChecked()
            self.distinct_roi = self.distinctAutoROI.isChecked()
            self.roi_hull = self.applyHullROI.isChecked()
            self.combine_roi = self.combineROI.isChecked()
            self.roi_tresh_type = "Global"
            self.tresh_size_roi = self.threshROISpinBox1.value()
            self.tresh_cst_roi = self.threshROISpinBox2.value()
            self.blur_size_roi = self.blurROISpinBox.value()
            self.bg_roi_apply = self.applybgROI.isChecked()
            self.bg_blur_size_roi = self.bgblurROISpinBox.value()
            self.bg_blend_roi = self.bgblendROISpinBox.value()
            self.framePos = 0
            self.roi_min_area = self.roiMinSpinBox.value()
            self.epsilon_fraction = 10**(self.epsilonSpinBox.value())
            self.roi_morph = self.morphROI.isChecked()
            self.x_roi_morph = self.morphXSpinBox.value()
            self.y_roi_morph = self.morphYSpinBox.value()
            self.calibFactor = self.lengthValue.value()/self.pixelValue.value()

            self.definePaths()
            
            self.seekSlider.blockSignals(True)
            self.seekSlider.setValue(0)
            self.seekSlider.setTickInterval(int(0.2 *self.frameCount))
            self.seekSlider.setSingleStep(int(0.1 *self.frameCount))
            self.seekSlider.setMaximum(int(self.frameCount))
            self.seekSlider.blockSignals(False)
            
            self.recordStatus = False
            
            self.analyzeVideo.blockSignals(True)
            self.analyzeVideo.setChecked(False)
            self.analyzeVideo.blockSignals(False)
            
            self.dftGroupBox.blockSignals(True)
            self.dftGroupBox.setChecked(False)
            self.dftGroupBox.blockSignals(False)

            self.bgGroupBox.blockSignals(True)
            self.bgGroupBox.setChecked(False)
            self.bgGroupBox.blockSignals(False)

            self.threshROIGroupBox.blockSignals(True)
            self.threshROIGroupBox.setChecked(False)
            self.threshROIGroupBox.blockSignals(False)

##            self.distinctAutoROI.setEnabled(False)
            self.roi_auto = False
            self.segment = False
            self.show_segment = self.showSegment.isChecked()
            self.show_fg = self.segmentFGButton.isChecked()
            self.show_bg = self.segmentBGButton.isChecked()
##            self.distinct_roi = False
            
##            self.distinctAutoROI.blockSignals(True)
##            self.distinctAutoROI.setChecked(False)
##            self.distinctAutoROI.blockSignals(False)

            self.correctZeroForce.blockSignals(True)
            self.correctZeroForce.setChecked(False)
            self.correctZeroForce.blockSignals(False)
            


            self.showContours.setEnabled(False)
            self.showEffect.setEnabled(False)
            
            
            self.forceData = ForceAnal()
            self.zeroForceData = ForceAnal() #initalise
            self.correctZeroForce.setEnabled(False)
            self.clearData.setEnabled(True)
            self.videoEffect.setCurrentIndex(0)
            self.videoEffect.model().item(5).setEnabled(False)
            self.videoEffect.model().item(6).setEnabled(True)

            self.videoFileNameLabel.setText("<b>Video file:</b>\n" + self.videoPath) #video path
            self.forceFileNameLabel.setText("Select force data from file menu")
            self.zeroForceFileNameLabel.setText("Select zero-force data from file menu")
##            self.statusBar.showMessage("Frame number: " + str(int(self.framePos)) + "\t("
##                                       + str(int((self.framePos)*100/self.frameCount)) +
##                                       "%)")
            self.statusBar.showMessage(str(int(self.frameCount)) + " frames loaded!")
            self.configRecWindow.videoTextbox.setText("")

            self.init_dict() #initialise dictionaries
            
            self.framePos = 1 #avoid array indexing issue
            self.frame_bin_roi_full = None

            self.curve1 = QScatterSeries()#initialise live plot curves
            self.curve2 = QScatterSeries()
            self.initialise_live_plot(self.curve1, Qt.blue) #contact area plot
            self.initialise_live_plot(self.curve2, Qt.red) # roi area plot

            self.msrmnt_num = None
            self.setWindowTitle(self.appVersion)

            if self.msrListMode == True: #continue loading force data etc
                self.load_measurement_data()

    def import_file_list(self): #import file list
        fileListPath, _ = QFileDialog.getOpenFileName(self, "Open File List")

        if fileListPath != "":
            self.folderPath = os.path.dirname(fileListPath)
            wb_obj = openpyxl.load_workbook(filename = fileListPath,
                                            read_only = True)# workbook object is created  
            sheet_obj = wb_obj.active

            m_row = sheet_obj.max_row

            self.measurmntList = []
            self.bottomviewList = []
            self.sideviewList = []
            self.forcedataList = []
             
            for i in range(2, m_row + 1): #save filename lists
                self.measurmntList.append(sheet_obj.cell(row = i, column = 1).value)
                self.bottomviewList.append(sheet_obj.cell(row = i, column = 2).value)
                self.sideviewList.append(sheet_obj.cell(row = i, column = 3).value)
                self.forcedataList.append(sheet_obj.cell(row = i, column = 4).value)

            wb_obj.close()
            self.chooseMsrmnt.setEnabled(True)
            self.statusBar.showMessage("Loaded file list from " + self.folderPath)
            print(self.measurmntList, self.folderPath)

    def choose_measurement(self): #choose measurement
        self.MsrDialog = QDialog(self)
        foldername = self.folderPath.split("/")[-1]
        self.MsrDialog.setWindowTitle("Measurement Picker: " + foldername)
        self.MsrDialog.resize(400, 500)
        gridLayout = QGridLayout(self.MsrDialog)
        self.listwidget = QListWidget(self.MsrDialog)
        okButton = QPushButton("Select", self.MsrDialog)
        okButton.clicked.connect(self.set_measurement)
        okButton.setDefault(True)
        
##        self.measurmntList = [1,2,3,4,5]
        itemlist = ["Measurement-" + str(x) for x  in self.measurmntList]
        self.listwidget.addItems(itemlist)

        gridLayout.addWidget(self.listwidget, 1, 0, 1, 1)
        gridLayout.addWidget(okButton, 2, 0, 1, 1)
        self.MsrDialog.show()

    def set_measurement(self): #assign corresponding filenames of the measurement
        self.MsrDialog.reject()
        self.msrListMode = True 
        self.msrmnt_num_current = int(self.listwidget.currentItem().text().split("-")[1])
      
        self.videoPath = self.folderPath + "/Imaging/Bottom View/" + \
                         self.bottomviewList[self.msrmnt_num_current-1]       
        self.load_video()

    def load_measurement_data(self): #load force data of measurement
        self.forceData = ForceAnal()
        self.forceData.force_filepath = self.folderPath +  \
                                        "/Force curves/" + \
                                        self.forcedataList[self.msrmnt_num_current-1]
        print(self.forceData.force_filepath)
        self.import_force_data()

        self.configRecWindow.videoTextbox.setText(self.folderPath + \
                                                  "/Imaging/Side View/" +
                                                  self.sideviewList[self.msrmnt_num_current-1])
        print("end")
        self.msrListMode = False
        self.msrmnt_num = self.msrmnt_num_current
        self.definePaths()
        self.setWindowTitle(self.appVersion + " Measurement-" + str(self.msrmnt_num))

    def loading_indicate(self, framenum): #loading
        self.statusBar.showMessage("Loading... Frame " + str(int(framenum)))
##        while(True):
##            ret, frame = cap.read()
##            if ret ==False:
##                self.frameCount = cap.get(cv2.CAP_PROP_POS_FRAMES)
##                print("framecount", self.frameCount)
##                cap.release()
##                break
        
        
    def playback(self): #set video playback status, Play: True / Pause: False
        if self.videoPath != "":
            print(self.playStatus)
            self.playStatus = not self.playStatus
            self.frameAction = "Play"
            self.playBtn.setIcon(QIcon('images/pause.png'))
            self.effectChain = [1, 1, 1, 1] #b/c, bg sub, filter, tresh
            print("play")
            while True:
                print("Loop start", time.time())
                
                self.ret, self.frame = self.cap.read()
                print("try", self.ret)
                if self.ret == False: #reset video at end or at error
                    print("if")
                    self.cap.release()
                    self.cap = cv2.VideoCapture(self.videoPath)
                    self.ret, self.frame = self.cap.read()
                    self.framePos = self.cap.get(cv2.CAP_PROP_POS_FRAMES)
                    self.playStatus = False #pause
                else:
                    self.framePos = self.cap.get(cv2.CAP_PROP_POS_FRAMES)
                    print("else", self.framePos)

                self.seekSlider.blockSignals(True)
                self.seekSlider.setValue(int(self.framePos))
                self.seekSlider.blockSignals(False)
                
                self.statusBar.showMessage("Frame number: " + str(int(self.framePos)) + "\t("
                                           + str(int((self.framePos)*100/self.frameCount)) +
                                           "%)" + "\tTime: " +
                                           "{0:.2f}".format(self.frameTime[int(self.framePos-1)]) + " s")
    
                print("frame no. get", self.framePos, self.frame.shape, time.time())

                
                self.frame_current = self.frame.copy()
                print("Frame copy", time.time())
                
                self.video_effect(self.frame) #apply filter, b/c, bg subtract effects
                print("Video Effect", time.time())
                self.roi_auto = self.threshROIGroupBox.isChecked()
                self.roi_hull = self.applyHullROI.isChecked()
                self.combine_roi = self.combineROI.isChecked()
                self.video_analysis() #tresholding and analysis
                print("Video Analysis", time.time())

##                if self.framePos == self.frameCount: #REMOVE -1 and CHECK
##                    self.cap.set(cv2.CAP_PROP_POS_FRAMES, 0)
##                    print("reset")
                print(self.playStatus, self.frameAction, self.framePos)

    ##            time.sleep(1/self.fpsSlider.value()) #playback speed
                cv2.waitKey(1)

                if self.playStatus == False: #pause
                    print("frame_urrent", self.frame_current.shape)
                    while True:
                        # print("x")
                        cv2.waitKey(1)
                        self.playBtn.setIcon(QIcon('images/play.png'))
                        if self.playStatus == True: #resume
                            self.playBtn.setIcon(QIcon('images/pause.png'))
                            break
                        if self.frameAction == "Next": #next frame
                            self.frameAction = "Pause"
                            print(self.cap.get(cv2.CAP_PROP_POS_FRAMES),
                                  self.framePos)
                            self.effectChain = [1, 1, 1, 1]
                            break
                        if self.frameAction == "Previous": #previous frame
                            #change below to consider last two frames as well.CHECK
    ##                        self.framePos = self.frameCount + self.framePos - 3 \
    ##                                   if self.framePos < 2 else self.framePos - 2
                            print(self.cap.get(cv2.CAP_PROP_POS_FRAMES),
                                  self.framePos)
                            self.framePos = self.frameCount - 1 \
                                       if self.framePos == 1 else self.framePos - 2
                            print(self.framePos)
                            self.cap.set(cv2.CAP_PROP_POS_FRAMES, self.framePos)
                            self.frameAction = "Pause"
                            print(self.cap.get(cv2.CAP_PROP_POS_FRAMES),
                                  self.framePos)
                            self.effectChain = [1, 1, 1, 1]
                            break
                        if self.frameAction == "Stop": #stop
                            print("stop")
                            break

                                        
                if self.frameAction == "Stop": #stop
                    #TO DO: make a video initialize function
                    print("stop 2")
                    self.playBtn.setIcon(QIcon('images/play.png'))
                    self.cap.set(cv2.CAP_PROP_POS_FRAMES, 0)
                    self.framePos = 1 #avoid array indexing issue (-1)
                    self.seekSlider.blockSignals(True)
                    self.seekSlider.setValue(0)
                    self.seekSlider.blockSignals(False)
                    self.ret, self.frame = self.cap.read()
                    self.frame_current = self.frame.copy()
##                    self.roiBound = [0, 0, self.frameWidth, self.frameHeight]
##                    self.video_effect(self.frame)
                    self.roi_auto = self.threshROIGroupBox.isChecked()
                    self.roi_hull = self.applyHullROI.isChecked()
                    self.combine_roi = self.combineROI.isChecked()
##                    self.video_analysis() #tresholding and analysis
                    
                    self.renderVideo("Raw", self.ret, self.frame)
                    self.cap.set(cv2.CAP_PROP_POS_FRAMES, 0)
                    self.playStatus = False
                    self.frame_current = self.frame.copy()
                    self.effectChain = [1, 1, 1, 1]
                    self.recordStatus = False

                    self.statusBar.showMessage("Frame number: " + str(int(self.framePos-1)) + "\t("
                                           + str(int((self.framePos-1)*100/self.frameCount)) +
                                           "%)" + "\tTime: " +
                                           "{0:.2f}".format(self.frameTime[int(self.framePos-1)]) + " s")
                    break

                self.frame = None

    def video_analysis(self):
        if self.videoPath != "":
            if self.analyzeVideo.isChecked() == True:
                print("video analysis")
                
                roi = self.roiBound
                self.frame_transformed = self.frame[roi[1]:roi[3], roi[0]:roi[2]].copy()
                if self.showEffect.isChecked() == True:
                    self.frame_contour = self.frame[roi[1]:roi[3], roi[0]:roi[2]].copy()
                else:
                    self.frame_contour = self.frame_current[roi[1]:roi[3], roi[0]:roi[2]].copy()
                    
                self.frame_bin, self.frame_masked, self.frame_contour, cpt, elp = \
                          self.getContours(self.threshType.currentText(),
                                           self.threshSpinBox1.value(),
                                           self.threshSpinBox2.value(),
                                           self.resizeROISpinBox.value(),
                                           self.segmentFGSpinBox.value(),
                                           self.segmentBGSpinBox.value(),
                                           self.minAreaFilter.value(),
                                           self.maxAreaFilter.value())

                #initialise dictionary
    ##            self.dataDict = {"Default" : 6 * [np.zeros(int(self.frameCount), np.float64)]}
                print("contour data length", len(self.contour_data[0]))

                #prevent data accumulation
                memory_max = 100000
                memory_current = len(self.contour_data[0])
                if memory_current > memory_max: 
                    self.clear_data()
                elif memory_current > int(0.9 * memory_max):
                    self.statusBar.showMessage("<b>WARNING! " +
                                               str(int(memory_current*100/memory_max)) +
                                               "% MEMORY USED!</b>")
    ##                cv2.waitKey(500)
                    
                #get contact area

                for k in self.roiDict.keys():
                    if len(self.roiDict.keys()) > 1 and k == "Default":
                        continue
    ##                contours = self.roiDict[k][2]

                    #use deepcopy to avoid referencing issue
                    data = copy.deepcopy(self.dataDict[k])
                    contactArea = copy.deepcopy(data[0])
                    contactLength = copy.deepcopy(data[1])
                    contourNumber = copy.deepcopy(data[2])
                    roiArea = copy.deepcopy(data[3])
                    roiLength = copy.deepcopy(data[4])
                    eccAvg = copy.deepcopy(data[5])
                    ellipse = copy.deepcopy(data[6])
                    # cpt = self.contourProperty(k, self.minAreaFilter.value(),
                    #                            self.maxAreaFilter.value())                
                    print(cpt[k])
                    contactArea[int(self.framePos-1)] = cpt[k][0]
                    # contactArea[int(self.framePos-1)] = np.count_nonzero(self.frame_masked==0)
                    contactLength[int(self.framePos-1)] = cpt[k][1]
                    contourNumber[int(self.framePos-1)] = cpt[k][2]
                    roiArea[int(self.framePos-1)] = cpt[k][3]
                    roiLength[int(self.framePos-1)] = cpt[k][4]
                    eccAvg[int(self.framePos-1)] = cpt[k][5]
                    ellipse[int(self.framePos-1)] = elp[k]
                    
                    #calibrate contact area
                    self.calibFactor = self.lengthValue.value()/self.pixelValue.value()
                    contactArea[int(self.framePos-1)] *= self.calibFactor**2
                    contactLength[int(self.framePos-1)] *= self.calibFactor
                    roiArea[int(self.framePos-1)] *= self.calibFactor**2
                    roiLength[int(self.framePos-1)] *= self.calibFactor

                    self.dataDict[k] = [contactArea, contactLength,
                                        contourNumber, roiArea,
                                        roiLength, eccAvg, ellipse]
                    
                    cv2.drawContours(self.frame_contour, [self.roiDict[k][3]],
                                     -1, (0,0,255), 2) #roi 
                    # cv2.drawContours(self.frame_contour, self.roiDict[k][2], -1,
                    #                  (0,255,0), -1) #contours
                    if self.roi_auto == True: 
                        cv2.drawContours(self.frame_contour, [self.roiDict[k][4]],
                                         -1, (255,0,0), 2)#auto roi
                
                # for k in self.roiDict.keys():
                #     if len(self.roiDict.keys()) > 1 and k == "Default":
                #         continue
                    # if len(self.roiDict[k][2]) > 0: # 1 point needed to stack
                    #     cont_comb = np.vstack(self.roiDict[k][2][i] for i in range(len(self.roiDict[k][2])))
                    #     if len(cont_comb) > 4: # min 5 points needed to fit ellipse
                    #         elps = cv2.fitEllipse(cont_comb) #fit and draw bounding ellipse
                    #         cv2.ellipse(self.frame_contour, elps,(51,255,255),2)
                    #     else:
                    #         elps = ((0,0),(0,0),0,1)
                    # else:
                    #     elps = ((0,0),(0,0),0,1)
                    # ellipse[int(self.framePos-1)] = list(elps) + [self.calibFactor] #ellipse center, axis, angle, length calibration
                    # self.dataDict[k][6] = ellipse

                    # cv2.drawContours(self.frame_contour, [self.roiDict[k][3]],
                    #                  -1, (0,0,255), 2) #roi 
                    # cv2.drawContours(self.frame_contour, self.roiDict[k][2], -1,
                    #                  (0,255,0), -1) #contours
                    # if self.roi_auto == True: 
                    #     cv2.drawContours(self.frame_contour, [self.roiDict[k][4]],
                    #                      -1, (255,0,0), 2)#auto roi
                
                self.effectChoices = {'Binary': self.frame_bin,
                                      'Masked': self.frame_masked,
                                      'Contours': self.frame_contour,
                                      'Transformed': self.frame_transformed,
                                      'Background': self.frameBackground,
                                      'Auto ROI': self.frame_bin_roi_full}
                print("effect choice")

                
                if self.videoEffect.currentText() == "Force/Area Plot":
                    self.forceData.getArea(self.frameTime, self.dataDict)
                    self.forceData.plotData(self.lengthUnit.currentText()) #prepare plot
                    self.w = self.roiBound[2] - self.roiBound[0]
                    self.h = self.roiBound[3] - self.roiBound[1]
                    dim = (1280, 1024) #CHECK!
    ##                dim = (2560, 1440)
                    self.frameEffect = cv2.resize(cv2.cvtColor(self.forceData.convertPlot(), cv2.COLOR_RGB2BGR),
                                                  dim, interpolation = cv2.INTER_AREA) 
                else:
                    self.frameEffect = self.effectChoices.get(self.videoEffect. \
                                                    currentText())
                    if self.effectContrast.isChecked() == True: #enhance effect contrast
                        frame_inv = 255 - self.frameEffect
                        max_intensity = np.max(frame_inv)
                        gain = int(255/max_intensity) if max_intensity != 0 else 1
                        self.frameEffect = 255 - (gain * frame_inv) #increase effect contrast
    ##            roi = self.roiBound
                if self.showContours.isChecked() == True:
    ##                if self.showEffect.isChecked() == True:
    ##                    frame_disp = self.frame[roi[1]:roi[3], roi[0]:roi[2]].copy()
    ##                else:
    ##                    frame_disp = self.frame_current[roi[1]:roi[3], roi[0]:roi[2]].copy()
    ##                
    ##                for k in self.roiDict.keys():
    ##                    if len(self.roiDict.keys()) > 1 and k == "Default":
    ##                        continue
    ##                    cv2.drawContours(frame_disp, [self.roiDict[k][3]],
    ##                                     -1, (0,0,255), 2)
    ##                    cv2.drawContours(frame_disp, self.roiDict[k][2], -1, (0,255,0), 1)
    ####                frame_disp = frame_disp[roi[1]:roi[3], roi[0]:roi[2]]
                    frame_disp = self.frame_contour.copy()
    ##                self.renderVideo("Raw", self.ret, frame_disp)
    ##                self.recordVideo(frame_disp, self.frameEffect)
                else:
                    if self.showEffect.isChecked() == True:
                        frame_disp = self.frame[roi[1]:roi[3], roi[0]:roi[2]].copy()
                    else:
                        frame_disp = self.frame_current[roi[1]:roi[3], roi[0]:roi[2]].copy()

                self.renderVideo("Raw", self.ret, frame_disp)
                self.renderVideo("Effect", self.ret, self.frameEffect)
                self.recordVideo(frame_disp, self.frameEffect)

                print("Render finish", time.time())
                print(self.framePos)

                self.statusBar.showMessage("Frame number: " + str(int(self.framePos)) + "\t("
                                           + str(int((self.framePos)*100/self.frameCount)) +
                                           "%)" + "\tTime: " +
                                           "{0:.2f}".format(self.frameTime[int(self.framePos-1)]) + " s")
                self.plot_live_data()
            else:
                roi = self.roiBound
                frame_disp = self.frame[roi[1]:roi[3], roi[0]:roi[2]].copy()
                self.renderVideo("Raw", self.ret, frame_disp)
                self.effectScene.removeItem(self.effectPixmapItem)
                self.effectPixmapItem = self.effectScene.addPixmap(self.blankPixmap)
                self.effectView.fitInView(self.rawPixmapItem, 1)
                self.showContours.setEnabled(False)


    def stop_video(self):
        if self.videoPath != "":
            self.frameAction = "Stop"
            roiCorners = np.array([[0, 0],[self.frameWidth, 0], 
                                        [self.frameWidth, self.frameHeight], 
                                        [0, self.frameHeight]],np.int32)
            self.roiBound = [0, 0, self.frameWidth, self.frameHeight]
            self.roiDict = {"Default": [roiCorners, self.roiBound, [],
                                        roiCorners, roiCorners]}
            self.init_dict() #initialise dictionaries
        
    def next_frame(self):
        if self.videoPath != "":
            self.frameAction = "Next"

    def previous_frame(self):
        if self.videoPath != "":
            self.frameAction = "Previous"

    def seek_video(self):
        if self.videoPath != "":
            self.cap.set(cv2.CAP_PROP_POS_FRAMES, self.seekSlider.value())
##            self.framePos = self.seekSlider.value()
##            if self.framePos == self.frameCount-1: #REMOVE -1 and CHECK
##                self.cap.set(cv2.CAP_PROP_POS_FRAMES, 0)

            self.ret, self.frame = self.cap.read()
            if self.seekSlider.value() == 0:
                self.cap.set(cv2.CAP_PROP_POS_FRAMES, 0)

            if self.ret == False: #CHECK
                print("if")
                self.cap.release()
                self.cap = cv2.VideoCapture(self.videoPath)
                self.ret, self.frame = self.cap.read()
##                self.framePos = self.cap.get(cv2.CAP_PROP_POS_FRAMES)
##            else:
            self.framePos = self.cap.get(cv2.CAP_PROP_POS_FRAMES)
            print("else", self.framePos)

            self.seekSlider.blockSignals(True)
            self.seekSlider.setValue(self.framePos)
            self.seekSlider.blockSignals(False)

            self.statusBar.showMessage("Frame number: " + str(int(self.framePos)) + "\t("
                                           + str(int((self.framePos)*100/self.frameCount)) +
                                           "%)" + "\tTime: " +
                                           "{0:.2f}".format(self.frameTime[int(self.framePos-1)]) + " s")
##            self.framePos = 1 if self.framePos == 0 else self.framePos #avoid array indexing issue
                
##            self.statusBar.showMessage("Frame number: " + str(int(self.framePos)) + "\t("
##                                       + str(int((self.framePos)*100/self.frameCount)) +
##                                       "%)")
            
                    
            
            self.frame_current = self.frame.copy()
            self.video_effect(self.frame_current) #apply filter, b/c, bg subtract effects
            self.roi_auto = self.threshROIGroupBox.isChecked()
            self.roi_hull = self.applyHullROI.isChecked()
            self.combine_roi = self.combineROI.isChecked()

            if self.framePos == 0:
##                self.blankPixmap = QPixmap('images/blank.png')
                self.statusBar.showMessage("Press play to begin")
                self.rawScene.removeItem(self.rawPixmapItem)
                self.rawPixmapItem = self.rawScene.addPixmap(self.blankPixmap)
                self.rawView.fitInView(self.rawPixmapItem, 1)
                self.effectScene.removeItem(self.effectPixmapItem)
                self.effectPixmapItem = self.effectScene.addPixmap(self.blankPixmap)
                self.effectView.fitInView(self.rawPixmapItem, 1)
            else:
                self.video_analysis() #tresholding and analysis


    def video_effect(self, frame): #apply effects in effects chain on self.frame
        print("video effect", self.effectChain)
        if self.effectChain[0] == 1: #1: brightness/contrast
            self.frame = self.applyBrightnessContrast(
                self.brightnessSlider.value(),self.contrastSlider.value(),
                frame)
            frame = self.frame
                
        if self.bgGroupBox.isChecked() == True and \
           self.effectChain[1] == 1: #2: background subtraction
##                self.subtract = True #CHECK
##                self.frameBackground = self.frameBackground #CHECK
            print(frame.shape, self.frameBackground.shape)
            if self.backgroundCorrection.currentText()== "Rolling Ball":
                frame_gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
                frame_corr, self.frameBackground = subtract_background_rolling_ball(frame_gray,
                                                                          self.bgSlider.value(),
                                                                     light_background=True,
                                                                     use_paraboloid=False,
                                                                     do_presmooth=True)
                self.frame = cv2.cvtColor(frame_corr, cv2.COLOR_GRAY2BGR)
            elif self.backgroundCorrection.currentText()== "Rolling Paraboloid":
                frame_gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
                frame_corr, self.frameBackground = subtract_background_rolling_ball(frame_gray,
                                                                          self.bgSlider.value(),
                                                                     light_background=True,
                                                                     use_paraboloid=True,
                                                                     do_presmooth=True)
                self.frame = cv2.cvtColor(frame_corr, cv2.COLOR_GRAY2BGR)
            elif self.backgroundCorrection.currentText()== "Proper Correction":
                self.frame = self.backgroundSubtract(frame,
                                                     self.frameBackground,
                                                     1)
            elif self.backgroundCorrection.currentText()== "Gaussian Correction":
                kernal = (self.bgSlider.value(), self.bgSlider.value())
                self.frameBackground = cv2.GaussianBlur(frame, kernal,0)
                self.frame = self.backgroundSubtract(frame, self.frameBackground,
                                                     self.bgAlphaSpinBox.value())
            elif self.backgroundCorrection.currentText()== "Average Correction":
                kernal = (self.bgSlider.value(), self.bgSlider.value())
                self.frameBackground = cv2.blur(frame, kernal)
                self.frame = self.backgroundSubtract(frame, self.frameBackground,
                                                     self.bgAlphaSpinBox.value())

            frame = self.frame
            
        if self.dftGroupBox.isChecked() == True and \
           self.effectChain[2] == 1: #3: dft filter
            if self.filterType.currentText() == "Fourier Filter":
                self.frame, _ = self.dftFilter(self.lowPassSlider.value(),
                                               self.highPassSlider.value(),
                                               frame)
            else:
                print("filter start", self.filterType.currentText())
                self.frame = self.imageFilter(self.filterType.currentText(),
                                              self.lowPassSlider.value(),
                                              self.highPassSlider.value(),
                                              frame)
            print("filtered", self.frame.shape)
##            frame = self.frame

    def threshold(self): #threshold change
        print("tresh", self.threshSlider1.value(), self.threshType)
        if self.threshType.currentText() == "Adaptive":
            if self.threshSlider1.value() %2 == 0: #make sure its odd
                self.threshSlider1.blockSignals(True)
                self.threshSlider1.setValue(self.threshSlider1.value() + 1)
                self.threshSlider1.blockSignals(False)
            if self.threshSpinBox1.value() %2 == 0: #make sure its odd
                self.threshSpinBox1.blockSignals(True)
                self.threshSpinBox1.setValue(self.threshSpinBox1.value() + 1)
                self.threshSpinBox1.blockSignals(False)

        changed_object = self.sender() #signal source
        if changed_object.__class__.__name__ == "QSlider":
            self.threshSpinBox1.blockSignals(True)
            self.threshSpinBox1.setValue(self.threshSlider1.value())
            self.threshSpinBox1.blockSignals(False)
            self.threshSpinBox2.blockSignals(True)
            self.threshSpinBox2.setValue(self.threshSlider2.value())
            self.threshSpinBox2.blockSignals(False)
        elif changed_object.__class__.__name__ == "QSpinBox":
            self.threshSlider1.blockSignals(True)
            self.threshSlider1.setValue(self.threshSpinBox1.value())
            self.threshSlider1.blockSignals(False)
            self.threshSlider2.blockSignals(True)
            self.threshSlider2.setValue(self.threshSpinBox2.value())
            self.threshSlider2.blockSignals(False)

        if len(self.frame) != 0:
            self.video_analysis()

    def segment_change(self): #image segment change
        if self.applySegment.isChecked() == True:
            self.segmentFGButton.setEnabled(True)
            self.segmentFGSlider.setEnabled(True)
            self.segmentFGSpinBox.setEnabled(True)
            self.segmentBGButton.setEnabled(True)
            self.segmentBGSlider.setEnabled(True)
            self.segmentBGSpinBox.setEnabled(True)
            self.showSegment.setEnabled(True)
            self.segment = True
        else:
            self.segmentFGButton.setEnabled(False)
            self.segmentFGSlider.setEnabled(False)
            self.segmentFGSpinBox.setEnabled(False)
            self.segmentBGButton.setEnabled(False)
            self.segmentBGSlider.setEnabled(False)
            self.segmentBGSpinBox.setEnabled(False)
            self.showSegment.setEnabled(False)
            self.segment = False
        
        if len(self.frame) != 0:
            self.video_analysis()        

    def segment_param_change(self): #segment parameter change
        changed_object = self.sender() #signal source
        if changed_object.__class__.__name__ == "QSlider":
            self.segmentFGSpinBox.blockSignals(True)
            self.segmentFGSpinBox.setValue(self.segmentFGSlider.value()/10)
            self.segmentFGSpinBox.blockSignals(False)
            self.segmentBGSpinBox.blockSignals(True)
            self.segmentBGSpinBox.setValue(self.segmentBGSlider.value())
            self.segmentBGSpinBox.blockSignals(False)          
        elif changed_object.__class__.__name__ in ["QSpinBox", "QDoubleSpinBox"]:
            self.segmentFGSlider.blockSignals(True)
            self.segmentFGSlider.setValue(int(self.segmentFGSpinBox.value()*10))
            self.segmentFGSlider.blockSignals(False)
            self.segmentBGSlider.blockSignals(True)
            self.segmentBGSlider.setValue(self.segmentBGSpinBox.value())
            self.segmentBGSlider.blockSignals(False)
        
        if len(self.frame) != 0:
            self.video_analysis()
            
    def segment_show(self): #show segmented image window
        self.show_segment = self.showSegment.isChecked()
        if len(self.frame) != 0:
            self.video_analysis()
    
    def segment_show_fg(self): #show sure foreground window
        self.show_fg = self.segmentFGButton.isChecked()
        if len(self.frame) != 0:
            self.video_analysis()
        
    def segment_show_bg(self): #show sure background window
        self.show_bg = self.segmentBGButton.isChecked()
        if len(self.frame) != 0:
            self.video_analysis()

    def dft_change(self): #dft filter change
        changed_object = self.sender() #signal source
        if changed_object.__class__.__name__ == "QSlider":
            self.lowPassSpinBox.blockSignals(True)
            self.lowPassSpinBox.setValue(self.lowPassSlider.value())
            self.lowPassSpinBox.blockSignals(False)
            self.highPassSpinBox.blockSignals(True)
            self.highPassSpinBox.setValue(self.highPassSlider.value())
            self.highPassSpinBox.blockSignals(False)
        elif changed_object.__class__.__name__ == "QSpinBox":
            self.lowPassSlider.blockSignals(True)
            self.lowPassSlider.setValue(self.lowPassSpinBox.value())
            self.lowPassSlider.blockSignals(False)
            self.highPassSlider.blockSignals(True)
            self.highPassSlider.setValue(self.highPassSpinBox.value())
            self.highPassSlider.blockSignals(False)

        self.filter_adjust()
        
        if len(self.frame) != 0:
            print(self.frame_current.size)
            if self.playStatus == False:
                if self.dftGroupBox.isChecked() == True:
                    self.effectChain = [1, 1, 1, 0] #order: b/c, bg sub, filter, tresh
                    self.video_effect(self.frame_current)
                    self.video_analysis()
                elif changed_object.__class__.__name__ == "QGroupBox":
                    self.frame = self.frame_current.copy()
                    self.effectChain = [1, 1, 0, 0]
                    self.video_effect(self.frame)
                    self.video_analysis()

    def filter_adjust(self): #adjust parameters of filtering
        ftypes = ['Gaussian Filter', 'Median Filter']
        if self.filterType.currentText() in ftypes:
            if self.lowPassSlider.value() < 3: #min value 3
                self.lowPassSlider.blockSignals(True)
                self.lowPassSlider.setValue(3)
                self.lowPassSlider.blockSignals(False)
            if self.lowPassSpinBox.value() < 3: #min value 3
                self.lowPassSpinBox.blockSignals(True)
                self.lowPassSpinBox.setValue(3)
                self.lowPassSpinBox.blockSignals(False)
            if self.lowPassSlider.value() %2 == 0: #make sure its odd
                self.lowPassSlider.blockSignals(True)
                self.lowPassSlider.setValue(self.lowPassSlider.value() + 1)
                self.lowPassSlider.blockSignals(False)
            if self.lowPassSpinBox.value() %2 == 0: #make sure its odd
                self.lowPassSpinBox.blockSignals(True)
                self.lowPassSpinBox.setValue(self.lowPassSpinBox.value() + 1)
                self.lowPassSpinBox.blockSignals(False)
        elif self.filterType.currentText() in ['Average Filter', 
                                               'Morph Open', 'Morph Close']:
            if self.lowPassSlider.value() < 1: #min value 1
                self.lowPassSlider.blockSignals(True)
                self.lowPassSlider.setValue(1)
                self.lowPassSlider.blockSignals(False)
            if self.lowPassSpinBox.value() < 1: #min value 1
                self.lowPassSpinBox.blockSignals(True)
                self.lowPassSpinBox.setValue(1)
                self.lowPassSpinBox.blockSignals(False)
            if self.highPassSlider.value() < 1: #min value 1
                self.highPassSlider.blockSignals(True)
                self.highPassSlider.setValue(1)
                self.highPassSlider.blockSignals(False)
            if self.highPassSpinBox.value() < 1: #min value 1
                self.highPassSpinBox.blockSignals(True)
                self.highPassSpinBox.setValue(1)
                self.highPassSpinBox.blockSignals(False)
        elif self.filterType.currentText() == 'Bilateral Filter':
            if self.highPassSlider.value() > 9: #max value 9
                self.highPassSlider.blockSignals(True)
                self.highPassSlider.setValue(9)
                self.highPassSlider.blockSignals(False)
            if self.highPassSpinBox.value() > 9: #max value 9
                self.highPassSpinBox.blockSignals(True)
                self.highPassSpinBox.setValue(9)
                self.highPassSpinBox.blockSignals(False)

    
    def bc_change(self): #brightness-contrast change
        print("bc_change")
        changed_object = self.sender() #signal source
        if changed_object.__class__.__name__ == "QSlider":
            self.brightnessSpinBox.blockSignals(True)
            self.brightnessSpinBox.setValue(self.brightnessSlider.value())
            self.brightnessSpinBox.blockSignals(False)
            self.contrastSpinBox.blockSignals(True)
            self.contrastSpinBox.setValue(self.contrastSlider.value())
            self.contrastSpinBox.blockSignals(False)
        elif changed_object.__class__.__name__ == "QSpinBox":
            self.brightnessSlider.blockSignals(True)
            self.brightnessSlider.setValue(self.brightnessSpinBox.value())
            self.brightnessSlider.blockSignals(False)
            self.contrastSlider.blockSignals(True)
            self.contrastSlider.setValue(self.contrastSpinBox.value())
            self.contrastSlider.blockSignals(False)
        if len(self.frame) != 0:
            if self.playStatus == False:
                print("bc anal")
                self.effectChain = [1, 1, 1, 1]
                self.video_effect(self.frame_current)
                self.video_analysis()

    def epsilon_change(self): #epsilon (roi hull) change
        changed_object = self.sender() #signal source
        if changed_object.__class__.__name__ == "QSlider":
            self.epsilonSpinBox.blockSignals(True)
            self.epsilonSpinBox.setValue(self.epsilonSlider.value())
            self.epsilonSpinBox.blockSignals(False)
        elif changed_object.__class__.__name__ == "QSpinBox":
            self.epsilonSlider.blockSignals(True)
            self.epsilonSlider.setValue(self.epsilonSpinBox.value())
            self.epsilonSlider.blockSignals(False)
        elif changed_object.__class__.__name__ == "QCheckBox":
            self.roi_hull = self.applyHullROI.isChecked()

        self.epsilon_fraction = 10**((self.epsilonSpinBox.value())/10)
        if len(self.frame) != 0 and self.roi_auto == True:
            self.video_analysis()

    def roi_morph_change(self): #change roi morph params
        if self.morphROI.isChecked() == True:
            self.morphXSpinBox.setEnabled(True)
            self.morphYSpinBox.setEnabled(True)
        else:
            self.morphXSpinBox.setEnabled(False)
            self.morphYSpinBox.setEnabled(False)
        self.roi_morph = self.morphROI.isChecked()
        self.x_roi_morph = self.morphXSpinBox.value()
        self.y_roi_morph = self.morphYSpinBox.value()
        if len(self.frame) != 0 and self.roi_auto == True:
            self.video_analysis()

    def roi_min_change(self): #change roi min
        self.roi_min_area = self.roiMinSpinBox.value()
        if len(self.frame) != 0 and self.roi_auto == True:
            self.video_analysis()

    def threshold_roi(self): #change threshold to detect roi (Wet case)
        self.roi_tresh_type = self.threshROIType.currentText()
        if self.threshROIType.currentText() == "Adaptive":
            self.threshROISlider1.setMinimum(3)
            self.threshROISpinBox1.setMinimum(3)
            if self.threshROISlider1.value() %2 == 0: #make sure its odd
                self.threshROISlider1.blockSignals(True)
                self.threshROISlider1.setValue(self.threshROISlider1.value() + 1)
                self.threshROISlider1.blockSignals(False)
            if self.threshROISpinBox1.value() %2 == 0: #make sure its odd
                self.threshROISpinBox1.blockSignals(True)
                self.threshROISpinBox1.setValue(self.threshROISpinBox1.value() + 1)
                self.threshROISpinBox1.blockSignals(False)
        else:
            self.threshROISlider1.setMinimum(1)
            self.threshROISpinBox1.setMinimum(1)

        changed_object = self.sender() #signal source

        if changed_object.__class__.__name__ == "QSlider":
            self.threshROISpinBox1.blockSignals(True)
            self.threshROISpinBox1.setValue(self.threshROISlider1.value())
            self.threshROISpinBox1.blockSignals(False)
            self.threshROISpinBox2.blockSignals(True)
            self.threshROISpinBox2.setValue(self.threshROISlider2.value())
            self.threshROISpinBox2.blockSignals(False)
        elif changed_object.__class__.__name__ == "QSpinBox":
            self.threshROISlider1.blockSignals(True)
            self.threshROISlider1.setValue(self.threshROISpinBox1.value())
            self.threshROISlider1.blockSignals(False)
            self.threshROISlider2.blockSignals(True)
            self.threshROISlider2.setValue(self.threshROISpinBox2.value())
            self.threshROISlider2.blockSignals(False)
            
        self.tresh_size_roi = self.threshROISpinBox1.value()
        self.tresh_cst_roi = self.threshROISpinBox2.value()
        
        if len(self.frame) != 0 and self.roi_auto == True:
            self.video_analysis()

    def blur_roi(self): #change roi  blur
        changed_object = self.sender() #signal source
        if changed_object.__class__.__name__ == "QSlider":
            self.blurROISpinBox.blockSignals(True)
            self.blurROISpinBox.setValue(self.blurROISlider.value())
            self.blurROISpinBox.blockSignals(False)
        elif changed_object.__class__.__name__ == "QSpinBox":
            self.blurROISlider.blockSignals(True)
            self.blurROISlider.setValue(self.blurROISpinBox.value())
            self.blurROISlider.blockSignals(False)

        self.blur_size_roi = self.blurROISpinBox.value()
        if len(self.frame) != 0 and self.roi_auto == True:
            self.video_analysis()        

    def bg_roi_change(self): #bg roi checkbox change
        changed_object = self.sender() #signal source
        if changed_object.__class__.__name__ == "QCheckBox":
            self.bg_roi_apply = self.applybgROI.isChecked()
            if self.bg_roi_apply == True:
                self.bgblurROISpinBox.setEnabled(True)
                self.bgblurROISlider.setEnabled(True)
                self.bgblendROISpinBox.setEnabled(True)
                self.bgblendROISlider.setEnabled(True)
            else:
                self.bgblurROISpinBox.setEnabled(False)
                self.bgblurROISlider.setEnabled(False)
                self.bgblendROISpinBox.setEnabled(False)
                self.bgblendROISlider.setEnabled(False)
        if len(self.frame) != 0 and self.roi_auto == True:
            self.video_analysis()

    def bg_blur_roi(self): # change bg roi blur
        changed_object = self.sender() #signal source
        if changed_object.__class__.__name__ == "QSlider":
            self.bgblurROISpinBox.blockSignals(True)
            self.bgblurROISpinBox.setValue(self.bgblurROISlider.value())
            self.bgblurROISpinBox.blockSignals(False)
        elif changed_object.__class__.__name__ == "QSpinBox":
            self.bgblurROISlider.blockSignals(True)
            self.bgblurROISlider.setValue(self.bgblurROISpinBox.value())
            self.bgblurROISlider.blockSignals(False)

        self.bg_blur_size_roi = self.bgblurROISpinBox.value()
        if len(self.frame) != 0 and self.roi_auto == True:
            self.video_analysis()        
        
    def bg_blend_roi(self): #change bg blend
        changed_object = self.sender() #signal source
        if changed_object.__class__.__name__ == "QSlider":
            self.bgblendROISpinBox.blockSignals(True)
            self.bgblendROISpinBox.setValue(self.bgblendROISlider.value()/100)
            self.bgblendROISpinBox.blockSignals(False)
        elif changed_object.__class__.__name__ == "QDoubleSpinBox":
            self.bgblendROISlider.blockSignals(True)
            self.bgblendROISlider.setValue(int(self.bgblendROISpinBox.value()*100))
            self.bgblendROISlider.blockSignals(False)

        self.bg_blend_roi = self.bgblendROISpinBox.value()
        if len(self.frame) != 0 and self.roi_auto == True:
            self.video_analysis()        
        

    def bgFrame(self):
        if len(self.frame) != 0:
            self.frameBackground = self.frame.copy()
            self.bgframeNumber = self.framePos
            self.statusBar.showMessage("Frame number " + str(self.bgframeNumber) + \
                                       " set as background frame")

    def bg_change(self): #background subtract change
        print("bg change")

        if self.backgroundCorrection.currentText() == "Gaussian Correction":
            if self.bgSlider.value() %2 == 0: #make sure its odd
                self.bgSlider.blockSignals(True)
                self.bgSlider.setValue(self.bgSlider.value() + 1)
                self.bgSlider.blockSignals(False)
            if self.bgSpinBox.value() %2 == 0: #make sure its odd
                self.bgSpinBox.blockSignals(True)
                self.bgSpinBox.setValue(self.bgSpinBox.value() + 1)
                self.bgSpinBox.blockSignals(False)
                
        changed_object = self.sender() #signal source

        if changed_object.__class__.__name__ == "QSlider":
            self.bgSpinBox.blockSignals(True)
            self.bgSpinBox.setValue(self.bgSlider.value())
            self.bgSpinBox.blockSignals(False)
        elif changed_object.__class__.__name__ == "QSpinBox":
            self.bgSlider.blockSignals(True)
            self.bgSlider.setValue(self.bgSpinBox.value())
            self.bgSlider.blockSignals(False)
            
        if len(self.frame) != 0:
            if self.playStatus == False:
                if self.bgGroupBox.isChecked() == True:
                    self.effectChain = [1, 1, 1, 0] #order: b/c, bg sub, filter, tresh
                    self.video_effect(self.frame_current)
                    self.video_analysis()
                elif changed_object.__class__.__name__ == "QGroupBox":
                    self.frame = self.frame_current.copy()
                    self.effectChain = [1, 0, 1, 0]
                    self.video_effect(self.frame)
                    self.video_analysis()

    def anal_change(self): #analysis video/ contour change
        print("anal change")
        if len(self.frame) != 0:
            if self.analyzeVideo.isChecked() == False:
                self.showContours.setEnabled(False)
                self.showEffect.setEnabled(False)
            else:
                self.showContours.setEnabled(True)
                self.showEffect.setEnabled(True)

            if self.playStatus == False:
                self.effectChain = [1, 1, 1, 0] #order: b/c, bg sub, filter, tresh
                self.video_effect(self.frame_current)
                self.video_analysis()
        else:
            self.analyzeVideo.setChecked(False)
    
    def effect_change(self): #effect dropdown change
        print("effect change")
        if len(self.frame) != 0:
            if self.playStatus == False and self.analyzeVideo.isChecked() == True:
                if self.videoEffect.currentText() == "Force/Area Plot":
                    self.forceData.getArea(self.frameTime, self.dataDict)
                    self.forceData.plotData(self.lengthUnit.currentText()) #prepare plot
                    self.w = self.roiBound[2] - self.roiBound[0]
                    self.h = self.roiBound[3] - self.roiBound[1]
                    dim = (1280, 1024) #CHECK!
                    self.frameEffect = cv2.resize(cv2.cvtColor(self.forceData.convertPlot(), cv2.COLOR_RGB2BGR),
                                                  dim, interpolation = cv2.INTER_AREA)

                else:
                    self.frameEffect = self.effectChoices.get(self.videoEffect. \
                                                    currentText())
                self.renderVideo("Effect", self.ret, self.frameEffect)

    def fps_change(self):#fps value change
        print("fps change", self.frameCount)
        self.frameRate = self.fpsSpinBox.value()
        self.frameTime = np.linspace(0,
                                     self.frameCount/self.frameRate,
                                     int(self.frameCount), dtype = np.float64)

    def auto_roi_change(self): #auto roi checkbox change
        self.roi_auto = self.threshROIGroupBox.isChecked()
        if self.roi_auto == True:
            self.videoEffect.model().item(5).setEnabled(True)
##            self.distinctAutoROI.setEnabled(True)
        else:
            self.videoEffect.model().item(5).setEnabled(False)
##            self.distinctAutoROI.blockSignals(True)
##            self.distinctAutoROI.setChecked(False)
##            self.distinct_roi = False
##            self.distinctAutoROI.blockSignals(False)
##            self.distinctAutoROI.setEnabled(False)
        if len(self.frame) != 0:# and self.roi_auto == True:
            self.video_analysis()

    def distinct_roi_change(self):
        self.distinct_roi = self.distinctAutoROI.isChecked()
        self.combine_roi = self.combineROI.isChecked()
        if len(self.frame) != 0 and self.roi_auto == True:
            self.video_analysis()

    #render video frame
    def renderVideo(self, view, ret, frame, roiCorners = np.array([],np.int32)): 

        if ret == True:
            
##            if view == "Raw" and self.analyzeVideo.isChecked() == False:
##                roi = self.roiBound
##                frame_view = frame[roi[1]:roi[3], roi[0]:roi[2]].copy()
##            else:
            frame_view = frame.copy()
            h, w = frame_view.shape[:2]
            
            if len(frame.shape) == 2: #binary image
                print("binary")
                byteValue = 1 * w
                img = QImage(frame_view, w, h, byteValue,
                       QImage.Format_Grayscale8)
            else:
                print("non-binary")
                byteValue = 3 * w
                cv2.cvtColor(frame_view, cv2.COLOR_BGR2RGB, frame_view)
##                cv2.polylines(frame_view, [roiCorners], True, (0,0,255), 2)          
                img = QImage(frame_view, w, h, byteValue,
                       QImage.Format_RGB888)
                        
            pixmap = QPixmap.fromImage(img)

            if view == "Raw":
                self.rawScene.removeItem(self.rawPixmapItem)
                self.rawScene.setSceneRect(0, 0, w, h)
                self.rawPixmapItem = self.rawScene.addPixmap(pixmap)
                self.rawView.fitInView(self.rawPixmapItem, 1)
                
            elif view == "Effect":
                self.effectScene.removeItem(self.effectPixmapItem)
                self.effectScene.setSceneRect(0, 0, w, h)
                self.effectPixmapItem = self.effectScene.addPixmap(pixmap)
                self.effectView.fitInView(self.effectPixmapItem, 1)

     
    def recordVideo(self, frame1, frame2):
        print("recordvideo")
        if self.recordStatus == True:
            h , w = 1024, 1280
            # dim = (w, h)
            if frame2.ndim == 2:
                frame2 = cv2.cvtColor(frame2, cv2.COLOR_GRAY2BGR)
            
##            if self.showContours.isChecked() == False:
##                roi = self.roiBound
##                self.merged_frame[:h, :w] = self.image_resize(frame1[roi[1]:roi[3],
##                                                              roi[0]:roi[2]],
##                                              w, h, inter = cv2.INTER_AREA)
##            else:
            self.merged_frame[:h, :w], scaleFactor = self.image_resize(frame1, w, h,
                                                                      inter = cv2.INTER_AREA)

            if self.configRecWindow.fourRec.isChecked() == True:
                if self.forceData.force_filepath == "" or self.cap2 == None:
                    root = Tk()
                    root.withdraw()
                    messagebox.showinfo("Error!", "Check 2nd video file or force data file. Not found!")
                    root.destroy()
                    self.record_frame() #finish recording
                    self.playStatus = False #pause video
                    return
##                frame2 = cv2.cvtColor(self.frame_contours, cv2.COLOR_GRAY2BGR)
                frame2 = self.frame_contour.copy()
                ret, frame3 = self.cap2.read()
                self.forceData.getArea(self.frameTime, self.dataDict)
                self.forceData.plotData(self.lengthUnit.currentText()) #prepare plot
                frame4 = cv2.resize(cv2.cvtColor(self.forceData.convertPlot(), cv2.COLOR_RGB2BGR),
                                                      (w, h), interpolation = cv2.INTER_AREA)
                
                if int(self.framePos) == self.forceData.plot_slice2.stop + 1:
                # if ret == False: #video at end
                    print("2nd video end")
                    self.cap2.release()
                    self.cap2 = None
                    self.record_frame() #finish recording
                    self.playStatus = True
                    return
                else:
                    framenumber1 = self.cap.get(cv2.CAP_PROP_POS_FRAMES)
                    framenumber2 = self.cap2.get(cv2.CAP_PROP_POS_FRAMES)
                    print("position", framenumber1, framenumber2)
                    if framenumber1 != framenumber2: #check both videos are in sync
                        root = Tk()
                        root.withdraw()
                        messagebox.showinfo("Error!", "Video frame numbers dont match!\n" +
                                            "Video-1 frame:\t" + str(framenumber1) + "\n" +
                                            "Video-2 frame:\t" + str(framenumber2))
                        root.destroy()
                        self.record_frame() #finish recording
                        self.playStatus = False #pause video
                        return
                    print("position", self.cap.get(cv2.CAP_PROP_POS_FRAMES),
                          self.cap2.get(cv2.CAP_PROP_POS_FRAMES))
                    
                    self.merged_frame[:h, w:], r = self.image_resize(frame3, w, h,
                                                              inter = cv2.INTER_AREA)
                    self.merged_frame[h:, :w], r = self.image_resize(frame2, w, h,
                                                              inter = cv2.INTER_AREA)
                    self.merged_frame[h:, w:], r = self.image_resize(frame4, w, h,
                                                              inter = cv2.INTER_AREA)
            else:
                self.merged_frame[:h, w:], r = self.image_resize(frame2, w, h,
                                                          inter = cv2.INTER_AREA)

            # Write time
            font = cv2.FONT_HERSHEY_SIMPLEX
            bottomLeftCornerOfText = (int(1.55*w), int(0.1*h))
            fontScale = 2
            fontColor = (0,200,200)
            thickness = 10
            lineType = 2
            print(self.frameTime.item(int(self.framePos-1)), bottomLeftCornerOfText)
            text = 'Time: ' + "{0:.3f}".format(self.frameTime.item(int(self.framePos-1))) + ' s'
            cv2.putText(self.merged_frame, text, 
                        bottomLeftCornerOfText, font,fontScale,
                        fontColor,thickness, lineType)

            #Draw scale bar
            print(scaleFactor, "scalef")
            pixLength = scaleFactor * self.pixelValue.value()
            scalepos1 = (int(0.8*w), int(0.95*h))
            scalepos2 = (int(scalepos1[0] + pixLength), scalepos1[1])
            scalelabelpos = (int(scalepos1[0] + 0.5 * (pixLength - 100)),
                             scalepos1[1] + 10) #length of label is 51 pixels
            cv2.line(self.merged_frame, scalepos1, scalepos2,
                     fontColor, thickness)
            fontScale = 1
            thickness = 5
            color = (0,200,200, 0)
            text = str(int(self.lengthValue.value())) + ' ' + self.lengthUnit.currentText()

            font = ImageFont.truetype("arial.ttf", 28, encoding="unic")
            img_pil = Image.fromarray(self.merged_frame)
            draw = ImageDraw.Draw(img_pil)
            draw.text(scalelabelpos, text, font = font, fill = color)
            self.merged_frame = np.array(img_pil)

            print(self.merged_frame.shape, w, h)
            self.out.write(self.merged_frame)

    def record_frame(self):
        print("record_frame")
        if self.recordStatus == True:
            self.out.release()
            self.recordBtn.setIcon(QIcon('images/record.png'))
            self.recordBtn.setEnabled(False)
            self.middleleftGroupBox.setEnabled(True)
            self.bcGroupBox.setEnabled(True)
            self.dftGroupBox.setEnabled(True)
            self.threshGroupBox.setEnabled(True)
            self.threshROIGroupBox.setEnabled(True)
            self.dataGroupBox.setEnabled(True)
            self.roiBtn.setEnabled(True)
            self.analyzeVideo.setEnabled(True)
            self.recordStatus = False
            self.playStatus = False
        else:
            self.recordBtn.setIcon(QIcon('images/recording.png'))
            self.middleleftGroupBox.setEnabled(False)
            self.bcGroupBox.setEnabled(False)
            self.dftGroupBox.setEnabled(False)
            self.threshGroupBox.setEnabled(False)
            self.threshROIGroupBox.setEnabled(False)
            self.dataGroupBox.setEnabled(False)
            self.roiBtn.setEnabled(False)
            self.analyzeVideo.setEnabled(False)
            self.recordStatus = True
            self.playback()
##        self.recordStatus = not self.recordStatus

    #function to resize frame for recording   
    def image_resize(self, image, width = None, height = None, inter = cv2.INTER_AREA):
        
        dim = None
        (h, w) = image.shape[:2]
        resized = np.zeros([height, width, 3], dtype = np.uint8)
        
        if width is None and height is None:
            return image, 0
        
        if width is None:
            r = height / float(h)
            dim = (int(w * r), height)
        elif height is None:
            r = width / float(w)
            dim = (width, int(h * r))

        else:
            rh = height / float(h)
            rw = width / float(w)

            if rh < rw:
                r = rh
                dim = (int(w * r), height)
            else:
                r = rw
                dim = (width, int(h * r))
                
        hdiff = int((height - dim[1])/2)
        wdiff = int((width - dim[0])/2)
 
        resized[hdiff:(hdiff + dim[1]),
                wdiff:(wdiff + dim[0])] = cv2.resize(image, dim,
                                                     interpolation = inter)
        print(dim, resized.shape)

        return resized, r

    def definePaths(self): #define filepaths
        if self.videoPath == "":
            self.dataPath = ""
            self.plotPath = ""
            self.recordingPath = ""
            self.summaryPath = ""
            self.contourDataPath = ""
        else:
            if self.msrmnt_num != None:
                    msrmnt_string = "Measurement-" + str(self.msrmnt_num) + " "
            else:
                msrmnt_string = ""
            analPath = os.path.dirname(os.path.dirname(os.path.dirname(self.videoPath))) \
                       + "/Analysis"
            self.dataPath = analPath + "/Data/" + msrmnt_string + \
                            self.videoPath.split('/')[-1][:-4] \
                            + "-data.txt"
            self.plotPath = analPath + "/Plots/" + msrmnt_string + \
                            self.videoPath.split('/')[-1][:-4] + "-plot.svg"
            self.recordingPath = analPath + "/Recording/" + msrmnt_string + \
                                 self.videoPath.split('/')[-1][:-4] + "-recording.avi"
            self.summaryPath = analPath + "/Summary/summary data.txt"
            self.contourDataPath = analPath + "/Contours/" + msrmnt_string + \
                            self.videoPath.split('/')[-1][:-4] + "-contour data.xlsx"

    def showPathWindow(self): #open path config window
        self.configPathWindow.dataPath.setText(self.dataPath)
        self.configPathWindow.plotPath.setText(self.plotPath)
        self.configPathWindow.recordingPath.setText(self.recordingPath)
        self.configPathWindow.summaryPath.setText(self.summaryPath)
        self.configPathWindow.contourDataPath.setText(self.contourDataPath)
        self.configPathWindow.showWindow()

    def setPaths(self): #set filepaths
        self.dataPath = self.configPathWindow.dataPath.toPlainText()
        self.plotPath = self.configPathWindow.plotPath.toPlainText()
        self.recordingPath = self.configPathWindow.recordingPath.toPlainText()
        self.summaryPath = self.configPathWindow.summaryPath.toPlainText()
        self.contourDataPath = self.configPathWindow.contourDataPath.toPlainText()
        self.configPathWindow.close()

    def showRecWindow(self): #open recording configuration window
        self.configRecWindow.showWindow(self.recordingPath)

    def configureRecord(self):
        if self.videoPath != "":
            self.w = self.roiBound[2] - self.roiBound[0]
            self.h = self.roiBound[3] - self.roiBound[1]
                   
            print("configurerecord", self.w, self.h)
            self.codecChoices = {'DIVX': cv2.VideoWriter_fourcc(*'DIVX'),
                                 'MJPG': cv2.VideoWriter_fourcc('M','J','P','G'),
                                      'FFV1': cv2.VideoWriter_fourcc('F','F','V','1')}
            fourcc = self.codecChoices.get(self.configRecWindow.codec.currentText())
            self.recordingPath = self.configRecWindow.textbox.toPlainText()
            if self.configRecWindow.fourRec.isChecked() == True:
                i = 2
            else:
                i = 1
            w = 2560
            h = i * 1024
            size = (w, h)
##            fps = self.frameRate
            fps = self.configRecWindow.fps.value() #fixed playback fps
            self.out = cv2.VideoWriter(self.recordingPath, fourcc, fps, size)

            self.merged_frame = np.empty([h, w, 3], dtype = np.uint8)
            print(self.recordingPath, self.merged_frame.shape)
            self.recordBtn.setEnabled(True)

        videofile2 = self.configRecWindow.videoTextbox.toPlainText() #second video
        print(videofile2)
        if videofile2 != "":
            self.cap2 = cv2.VideoCapture(videofile2)
        self.configRecWindow.close()
        self.seekSlider.setValue(0) #reset to beginning
        self.showContours.setChecked(False) #uncheck show contours
        self.clear_data() #clear data
        
    def configurePlot(self): #set plot design flags
        self.plotSequence()
        self.configPlotWindow.close()

    def updatePlot(self): #update plot
        if self.videoPath != "":
            if self.forceData.force_filepath != "":
                self.init_plotconfig()
                if self.forceData.ptsnumber != 0: #when no force file loaded
                    self.forceData.calcData()
                self.effect_change()
            self.plot_live_data()

    def plotSequence(self): #sequence of plot calculations
        if self.forceData.force_filepath != "":
            self.forceData.dataClean()
            self.defl_vert1_raw = self.forceData.defl_vert1.copy() #copy of raw vert data
        self.zero_force_correct()
        self.updatePlot()

    def update_calib(self): #update area data on calibration
        if self.videoPath != "":
            for k in self.roiDict.keys():
                if len(self.roiDict.keys()) > 1 and k == "Default":
                    continue
                # calibFactorOld = self.calibFactor
                self.calibFactor = self.lengthValue.value()/self.pixelValue.value()
                self.dataDict[k][0][int(self.framePos-1)] *= self.calibFactor**2
                self.dataDict[k][1][int(self.framePos-1)] *= self.calibFactor
                self.dataDict[k][3][int(self.framePos-1)] *= self.calibFactor**2
                self.dataDict[k][4][int(self.framePos-1)] *= self.calibFactor
            self.plotSequence()
    
    def init_plotconfig(self): #initialize plot configuration window
        self.forceData.flag_ca = self.configPlotWindow.showContactArea.isChecked()
        self.forceData.flag_ra = self.configPlotWindow.showROIArea.isChecked()
        self.forceData.flag_cl = self.configPlotWindow.showContactLength.isChecked()
        self.forceData.flag_rl = self.configPlotWindow.showROILength.isChecked()
        self.forceData.flag_cn = self.configPlotWindow.showContactNumber.isChecked()
        self.forceData.flag_ecc = self.configPlotWindow.showEcc.isChecked()
        self.forceData.flag_lf = self.configPlotWindow.showLateralForce.isChecked()
        self.forceData.flag_zp = self.configPlotWindow.showZPiezo.isChecked()
        self.forceData.flag_xp = self.configPlotWindow.showXPiezo.isChecked()
        self.forceData.flag_ap = self.configPlotWindow.showAdhesion.isChecked()
        self.forceData.flag_fp = self.configPlotWindow.showFriction.isChecked()
        self.forceData.flag_st = self.configPlotWindow.showStress.isChecked()
        self.forceData.flag_zd = self.configPlotWindow.showDeformation.isChecked()
        self.forceData.x_var = self.configPlotWindow.xAxisParam.currentText()
        self.forceData.calib_lat1 = self.configPlotWindow.latCalibEq.text()
        self.forceData.invert_latf = self.configPlotWindow.invertLatForce.isChecked()
        self.forceData.flag_ct = self.configPlotWindow.applyCrossTalk.isChecked()
        self.forceData.ctv_slope = self.configPlotWindow.vertCrossTalk.value()
        self.forceData.ctl_slope = self.configPlotWindow.latCrossTalk.value()
##        self.forceData.zero_stop = self.configPlotWindow.zeroRange2.value()
##        self.forceData.adh_start = self.configPlotWindow.adhRange1.value()
##        self.forceData.adh_stop = self.configPlotWindow.adhRange2.value()
##        self.forceData.prl_start = self.configPlotWindow.prlRange1.value()
##        self.forceData.prl_stop = self.configPlotWindow.prlRange2.value()
        self.forceData.rangeDict = self.configPlotWindow.rangeDict
        self.forceData.flag_zshift = self.configPlotWindow.zeroShift.isChecked()
        self.forceData.flag_lf_filter = self.configPlotWindow.filterLatF.isChecked()
        self.forceData.window_length = self.configPlotWindow.filter_wind.value()
        self.forceData.polyorder = self.configPlotWindow.filter_poly.value()
        self.forceData.startFull = self.configPlotWindow.startFull.value()
        self.forceData.endFull = self.configPlotWindow.endFull.value()
        self.forceData.noiseSteps = self.configPlotWindow.noiseSteps.text()
        self.forceData.legendPos = self.configPlotWindow.legendPos.text()
        self.forceData.flag_fit = self.configPlotWindow.fittingGroupBox.isChecked()
        self.forceData.fit_x = self.configPlotWindow.xFit.currentText()
        self.forceData.fit_y = self.configPlotWindow.yFit.currentText()
        self.forceData.startFit = self.configPlotWindow.fitStart.value()
        self.forceData.endFit = self.configPlotWindow.fitStop.value()
        self.forceData.fit_pos = self.configPlotWindow.fitPos.text()
        self.forceData.fit_show = self.configPlotWindow.showFitEq.isChecked()
        self.forceData.k_beam = self.configPlotWindow.kBeam.text()
        self.forceData.deform_tol = self.configPlotWindow.deformStart.value()
        
    def roiDraw(self): #draw roi
        if len(self.frame) != 0:
            print(self.frame.shape)
##            self.roiCorners = np.array([],np.int32)
            frame_dup = self.frame_current.copy()
##            cv2.polylines(frame_dup, [self.roiCorners],
##                      True, (0,0,255), 2) #final polygon roi
            roiCorners = draw_roi.roi_dry("Frame", frame_dup)
            print("roiCorners")
            roiBound = self.roiBoundingRectangle(roiCorners) #(xmin, ymin, xmax, ymax)
##            roiCorners = roiCorners - [roiBound[0],
##                                                 roiBound[1]]
            
##            self.video_effect(self.frame_current)
##            self.video_analysis() #CHECK

            #reinitialise draw_roi module variables
            draw_roi.trigger = 0 
            draw_roi.pts = []
            return roiCorners, roiBound

    def roiBoundingRectangle(self, roiCorners): #get bounding rectangle
        xmin = min(roiCorners[:, 0])
        xmax = max(roiCorners[:, 0])
        ymin = min(roiCorners[:, 1])
        ymax = max(roiCorners[:, 1])
        roi = [xmin, ymin, xmax, ymax]
        return roi

    def roiMerge(self): #combine multiple ROIs
        key = self.configROIWindow.roiNum.value()
        label = self.configROIWindow.roiDict[key]
        rc, rb = self.roiDraw()
        self.roiDict[label] = [rc, rb, [], [], []]
        
        self.getRoiBound()
        for k in self.roiDict.keys():
            if len(self.roiDict.keys()) > 1 and k == "Default":
                continue
            self.roiDict[k][3] = self.roiDict[k][0] - [self.roiBound[0],
                                                       self.roiBound[1]]
            self.dataDict[k] = 6 * [np.zeros(int(self.frameCount), np.float64)] + \
                               [[[(0,0),(0,0),0,1]]*int(self.frameCount)]
        self.contour_data = [[], [], [], [], [], [], [], []]
        self.video_effect(self.frame_current)
        self.video_analysis() #CHECK

    def getRoiBound(self): #get roiBound
        xmin, ymin, xmax, ymax = [], [], [], []
        for k in self.roiDict.keys():
            if len(self.roiDict.keys()) > 1 and k == "Default":
                continue
            xmin.append(self.roiDict[k][1][0])
            ymin.append(self.roiDict[k][1][1])
            xmax.append(self.roiDict[k][1][2])
            ymax.append(self.roiDict[k][1][3])
        self.roiBound = [min(xmin), min(ymin), max(xmax), max(ymax)]

    def closeROIWindow(self): #close roi window
        print("x")
        self.configPlotWindow.roiChoice.blockSignals(True)
        self.configPlotWindow.roiChoice.clear()
        self.configPlotWindow.roiChoice.addItem("Default")

        self.configPlotWindow.rangeDict = {"Default" : [[0,1],[0,100],[0,100],
                                                        [0,100],[0,100],[0,1]]}
        self.dataDict = {"Default" : 6 * [np.zeros(int(self.frameCount), np.float64)] + \
                         [[[(0,0),(0,0),0,1]]*int(self.frameCount)]}
        for k in self.configROIWindow.roiDict.values(): #update dictionary and combobox
            self.configPlotWindow.roiChoice.addItem(k)
            self.configPlotWindow.rangeDict[k] = [[0,1],[0,100],[0,100],
                                                  [0,100],[0,100], [0,1]]
            self.dataDict[k] = 6 * [np.zeros(int(self.frameCount), np.float64)] + \
                               [[[(0,0),(0,0),0,1]]*int(self.frameCount)]

        self.contour_data = [[], [], [], [], [], [], [], []]

        keys = list(self.roiDict.keys())
        for k in keys: #delete non-existant keys from roiDict
            if k == "Default":
                continue
            if k not in self.configROIWindow.roiDict.values():
                del self.roiDict[k]

        if len(self.roiDict.keys()) > 1: #set to first roi
            self.configPlotWindow.roiChoice.setCurrentIndex(1)
        self.configPlotWindow.roiChoice.blockSignals(False)

        self.getRoiBound() #update roi bounds
        for k in self.roiDict.keys():
            if len(self.roiDict.keys()) > 1 and k == "Default":
                continue
            self.roiDict[k][3] = self.roiDict[k][0] - [self.roiBound[0],
                                                       self.roiBound[1]]
            self.dataDict[k] = 6 * [np.zeros(int(self.frameCount), np.float64)] + \
                               [[[(0,0),(0,0),0,1]]*int(self.frameCount)]

        self.plotSequence()
        self.video_effect(self.frame_current)
        self.video_analysis()
        self.configROIWindow.close()

    def init_dict(self): #initialise roi dictionaries/roi labels
        self.configPlotWindow.roiChoice.blockSignals(True)
        self.configPlotWindow.roiChoice.clear()
        self.configPlotWindow.roiChoice.addItem("Default")
        self.configPlotWindow.roiChoice.blockSignals(False)
        self.configPlotWindow.rangeDict = {"Default" : [[0,1],[0,100],
                                                        [0,100],[0,100],
                                                        [0,100],[0,1]]}

        self.configROIWindow.roiDict = {}
        self.configROIWindow.roiDef.setText("ROI Definition:")
        self.configROIWindow.roiNum.blockSignals(True)
        self.configROIWindow.roiNum.setValue(1)
        self.configROIWindow.roiNum.blockSignals(False)
        self.configROIWindow.roiLabel.blockSignals(True)
        self.configROIWindow.roiLabel.setText("")
        self.configROIWindow.roiLabel.blockSignals(False)

        
    def measureScale(self): #measure length in pixels
        if len(self.frame) != 0:
            window_name = "Drag left mouse key and release. Right click: delete. Enter: Continue"
            img_disp = self.frame.copy()
            img_disp_clear = self.frame.copy()
            cv2.namedWindow(window_name,cv2.WINDOW_KEEPRATIO)
            cv2.moveWindow(window_name, 0, 0)
            global trigger, click
            pts = []
            trigger = 0
            click = 0
            
            def callback(event, x, y, flags, param):

                global trigger, click

                if  event == cv2.EVENT_LBUTTONDOWN:
                    pts.append([x,y])
                    click = 1
                elif event == cv2.EVENT_LBUTTONUP and click == 1:
                    click = 0
                    pts.append([x,y])
                    cv2.line(img_disp_clear, tuple(pts[0]),tuple(pts[1]), 
                                               (0,0,255), 2)
                    length = np.linalg.norm(np.array(pts[0])-np.array(pts[1]))
                    angle = np.rad2deg(np.arctan((pts[0][1] - pts[1][1])/
                                                 (pts[1][0] - pts[0][0])))
                    print("length1", length)
                    font = cv2.FONT_HERSHEY_SIMPLEX
                    posx = int((pts[0][0]+pts[1][0])/2)
                    posy = int((pts[0][1]+pts[1][1])/2)
                    cv2.putText(img_disp_clear, "{0:.2f}".format(length) + ' px  '
                                + "{0:.2f}".format(angle) +' deg', (posx, posy), font,
                                1,(0,255,0),2)
                elif event == cv2.EVENT_RBUTTONDOWN:
                    pts.clear()
                    trigger = 1
                    print("pp")
                else:
                    if len(pts) == 1:
                        cv2.line(img_disp, tuple(pts[0]), tuple([x,y]), 
                                               (0,0,255), 1)
                    
            cv2.setMouseCallback(window_name, callback)
            
            while True:           
                cv2.imshow(window_name,img_disp)
                if trigger == 1:
                    print("trigger")
                    img_disp_clear = self.frame.copy()
                    trigger = 0
                    
                img_disp = img_disp_clear.copy()
                
                key = cv2.waitKey(10) & 0xFF
            
                if key == 13: #press enter to continue
                    cv2.destroyAllWindows()
                    print("pts", pts, trigger)
                    if pts:
                        length = np.linalg.norm(np.array(pts[0])-np.array(pts[1]))
                        print(length)
                        self.pixelValue.setValue(round(length,2))
                    break
    ##                return pts
        

    def plot_live_data(self): #plot graph
        if self.forceData.force_filepath == "": #area plot only           
            if self.frameCount != 1: #ignore for image
                self.plotWindow.liveChart.removeAllSeries()
                self.plotWindow.liveChartScene.removeItem(self.plotWindow.liveChart)
                self.plotWindow.liveChartView.resetCachedContent()
                self.plotWindow.liveChart = QChart() #live chart
                self.plotWindow.liveChart.setSizePolicy(QSizePolicy.Minimum, QSizePolicy.Minimum)
                self.plotWindow.liveChart.legend().hide()
                print(self.plotWindow.liveChartView.size())

                w, h = (self.plotWindow.liveChartView.size().width(),
                                                          self.plotWindow.liveChartView.size().height())
                print(w,h)
                self.plotWindow.liveChart.setMinimumSize(w, h)
            
                self.curve1 = QScatterSeries()#initialise live plot curves
                self.initialise_live_plot(self.curve1, Qt.blue) #contact area plot
                print(self.frameTime.shape)
                for k in self.roiDict.keys():
                    if len(self.roiDict.keys()) > 1 and k == "Default":
                        continue
                    self.curve1.append(self.series_to_polyline(self.frameTime,
                                                              self.dataDict[k][0]))
                self.plotWindow.liveChart.addSeries(self.curve1)
                if self.roi_auto == True: # roi area plot
                    self.curve2 = QScatterSeries()
                    self.initialise_live_plot(self.curve2, Qt.red)
                    for k in self.roiDict.keys():
                        if len(self.roiDict.keys()) > 1 and k == "Default":
                            continue
                            self.curve2.append(self.series_to_polyline(self.frameTime,
                                                                      self.dataDict[k][3]))
                    self.plotWindow.liveChart.addSeries(self.curve2)
                
                self.plotWindow.liveChart.createDefaultAxes()
                self.plotWindow.liveChartScene.addItem(self.plotWindow.liveChart)
                self.plotWindow.liveChartView.setScene(self.plotWindow.liveChartScene)
                print("live plot end")
        elif self.forceData.fig1_close == False: #area and force plot
            self.forceData.getArea(self.frameTime, self.dataDict)
            self.forceData.plotData(self.lengthUnit.currentText())
            self.forceData.showPlot()
        
    def plot_data(self): #plot graph
        if self.forceData.force_filepath == "":
            self.plotWindow.home() #live area data show
        else:
            if self.playStatus == True: #pause video if running (bug)
                self.playBtn.click() 
            self.forceData.fig1_close = False #area data show
            self.forceData.getArea(self.frameTime, self.dataDict)
            self.forceData.plotData(self.lengthUnit.currentText())
            self.forceData.showPlot()
            plt.show()
        
    def initialise_live_plot(self, curve, color): #initalise live plot
        pen = curve.pen()
        pen.setColor(color)#Qt.blue
        pen.setWidthF(1)
        curve.setPen(pen)
        curve.setUseOpenGL(True)
        curve.setMarkerSize(4.0)    

    def series_to_polyline(self, xdata, ydata): #convert plot data for Qt
        """Convert series data to QPolygon(F) polyline
        This code is derived from PythonQwt's function named 
        `qwt.plot_curve.series_to_polyline`"""
        xsize = len(xdata)
        ysize = len(ydata)
        if xsize != ysize:
            root = Tk()
            root.withdraw()
            messagebox.showinfo("Live Plot Error!", "Check force file/video file\n" + \
                                "Exception: x axis and y axis array sizes don't match")
            root.destroy()
            self.playStatus = False
            
        polyline = QPolygonF(xsize)
        pointer = polyline.data()
        dtype, tinfo = np.float, np.finfo  # integers: = np.int, np.iinfo
        pointer.setsize(2*polyline.size()*tinfo(dtype).dtype.itemsize)
        memory = np.frombuffer(pointer, dtype)
        memory[:(xsize-1)*2+1:2] = xdata
        memory[1:(ysize-1)*2+2:2] = ydata
        return polyline   
        
    
    def save_data(self): #save data and plots
        if len(self.frame) != 0:
            roi_corners = []
            roi_labels = []
            dataDict_actual = self.dataDict.copy()
            for k in self.dataDict.keys():
                if len(self.dataDict.keys()) > 1 and k == "Default":
                    del dataDict_actual["Default"]
                    continue
                roi_corners.append(self.roiDict[k][0].tolist())
                roi_labels.append(k)
           
            if self.configPathWindow.dataGroupBox.isChecked() == True:
                data = np.stack(np.array(list(dataDict_actual.values())).T, axis = 1)
                data_string = '\n'.join([''.join("{0:.8f}".format(x)+ '\t' +str(y) +
                                        '\t' + str(z) + '\t' + str((w)) + '\t' +
                                        str(v) + '\t' + str(u) +
                                        '\t' + str(t)) \
                                        for x, y, z, w, v, u, t in zip(self.frameTime, \
                                                         data[0], data[1], data[2], \
                                                         data[3], data[4], data[5])]) 


                with open(self.dataPath, "w", encoding = "utf_8") as f: #save area data
                    f.write(self.appVersion + "\n")
                    f.write("Date\t" + time.strftime("%d/%m/%Y, %H:%M:%S") + "\n")
                    f.write("Measurement number\t" + str(self.msrmnt_num) + "\n")
                    f.write("Video file\t" + self.videoPath + "\n")
                    f.write("Threshold Type\t" + str(self.threshType.currentText()) + "\n")
                    f.write("Threshold Size\t" + str(self.threshSpinBox1.value()) + "\n")
                    f.write("Threshold Constant\t" + str(self.threshSpinBox2.value()) + "\n")
                    f.write("Apply Segment\t" + str(self.applySegment.isChecked()) + "\n")
                    f.write("Segment FG\t" + str(self.segmentFGSpinBox.value()) + "\n")
                    f.write("Segment BG\t" + str(self.segmentBGSpinBox.value()) + "\n")
                    f.write("Auto detect ROI\t" + str(self.threshROIGroupBox.isChecked()) + "\n")
                    f.write("Distinct ROI\t" + str(self.distinctAutoROI.isChecked()) + "\n")
                    f.write("Combine ROI\t" + str(self.combineROI.isChecked()) + "\n")
                    f.write("Apply Hull\t" + str(self.applyHullROI.isChecked()) + "\n")
                    f.write("ROI Minimum\t" + str(self.roiMinSpinBox.value()) + "\n")
                    f.write("Resolution\t" + str(self.epsilonSpinBox.value()) + "\n")
                    f.write("Resize factor\t" + str(self.resizeROISpinBox.value()) + "\n")
                    f.write("Apply Morph\t" + str(self.morphROI.isChecked()) + "\n")
                    f.write("Morph Kernal\t" + str([self.morphXSpinBox.value(),self.morphYSpinBox.value()]) + "\n")
                    f.write("Apply ROI BG Correction\t" + str(self.applybgROI.isChecked()) + "\n")
                    f.write("ROI BG Correction Value\t" + str(self.bgblurROISpinBox.value()) + "\n")
                    f.write("ROI BG Blend fraction\t" + str(self.bgblendROISpinBox.value()) + "\n")
                    f.write("ROI Threshold Type\t" + str(self.threshType.currentText()) + "\n")
                    f.write("ROI Threshold Size\t" + str(self.threshROISpinBox1.value()) + "\n")
                    f.write("ROI Threshold Constant\t" + str(self.threshROISpinBox2.value()) + "\n")
                    f.write("ROI Blur Size\t" + str(self.blurROISpinBox.value()) + "\n")
                    f.write("Brightness\t" + str(self.brightnessSpinBox.value()) + "\n")
                    f.write("Contrast\t" +  str(self.contrastSpinBox.value()) + "\n")
                    f.write("Minimum Area\t" + str(self.minAreaFilter.value()) + "\n")
                    f.write("Maximum Area\t" + str(self.maxAreaFilter.value()) + "\n")
                    f.write("Apply Filter\t" + str(self.dftGroupBox.isChecked()) + "\n")
                    f.write("Filter Type\t" + str(self.filterType.currentText()) + "\n")
                    f.write("Filter Param1\t" +  str(self.lowPassSpinBox.value()) + "\n")
                    f.write("Filter Param2\t" + str(self.highPassSpinBox.value()) + "\n")
                    f.write("Subtract Background\t" + str(self.bgGroupBox.isChecked()) + "\n")
                    f.write("BG Correction\t" + str(self.backgroundCorrection.currentText()) + "\n")
                    f.write("BG Correction Value\t" + str(self.bgSpinBox.value()) + "\n")
                    f.write("BG Blend fraction\t" + str(self.bgAlphaSpinBox.value()) + "\n")
                    f.write("BG Frame no.\t" + str(self.bgframeNumber) + "\n")
                    f.write("Frames per second\t" + str(self.fpsSpinBox.value()) + "\n")
                    f.write("ROI Corners \t" + str(roi_corners) + "\n")
                    print("Calibration factor [" + self.lengthUnit.currentText() + "/px]\t" + str(self.calibFactor) + "\n")
                    f.write("Calibration factor [" + self.lengthUnit.currentText() + "/px]\t" +
                                "{0:.8f}".format(self.calibFactor) + "\n")
                    f.write("Noise filter window\t" +  str(self.configPlotWindow.filter_wind.value()) + "\n")
                    f.write("Noise filter polyord\t" + str(self.configPlotWindow.filter_poly.value()) + "\n")
                    f.write("ROI Labels\t" + str(roi_labels) + "\n")
                    f.write("Lateral Force Calibration [μN]\t" +
                            self.configPlotWindow.latCalibEq.text() + "\n")
                    f.write("Apply Cross Talk \t" +
                            str(self.configPlotWindow.applyCrossTalk.isChecked()) + "\n")
                    f.write("Vertical Cross Talk [μN/μN]\t" +
                            str(self.configPlotWindow.vertCrossTalk.value()) + "\n")
                    f.write("Lateral Cross Talk [μN/μN]\t" +
                            str(self.configPlotWindow.latCrossTalk.value()) + "\n")
                    f.write("Apply Zero-Force Correction\t" +
                            str(self.correctZeroForce.isChecked()) + "\n")
                    
                    f.write("Time [s]\tContact_Area [" + self.lengthUnit.currentText() + "2]\t" +
                            "Contact_Length [" + self.lengthUnit.currentText() + "]\t" + 
                            "Contour_Number\tROI_Area [" + self.lengthUnit.currentText() + "2]\t" +
                            "ROI_Length [" + self.lengthUnit.currentText() + "]\tMedian_Eccentricity\n")
                    f.write(data_string)
                print(self.forceData)
            
            if self.forceData.force_filepath != "":
                self.forceData.getArea(self.frameTime, self.dataDict)
                if self.configPathWindow.plotGroupBox.isChecked() == True:
                    self.forceData.plotData(self.lengthUnit.currentText()) #prepare plot
                    self.forceData.savePlot(self.plotPath) #save force-area plot
                if self.configPathWindow.summaryGroupBox.isChecked() == True:
                    videofile1 = self.videoPath.split('/')[-1][:-4]
                    videofile2 = self.configRecWindow.videoTextbox.toPlainText().split('/')[-1][:-4]
                    print("video", videofile2, videofile1)
                    zeroforcefile = self.zeroForceData.force_filepath if \
                                    self.correctZeroForce.isChecked()== True else ""
                    self.forceData.saveSummaryData(videofile1, videofile2, zeroforcefile,
                                                   self.lengthUnit.currentText(),
                                                   self.msrmnt_num, self.summaryPath) #save summary data

            if self.configPathWindow.contourGroupBox.isChecked() == True:
                #save contour data
                contour_header = ["Frame_no", "ROI_Label", "Contour_ID", "Area",
                                  "Length", "Eccentricity", "Moments", "Array"]
                contour_datadict = dict(zip(contour_header, self.contour_data))
                self.contour_data = [[], [], [], [], [], [], [], []]
                df_contour = pd.DataFrame(contour_datadict)
                del contour_datadict
                #process contour data in separate thread
                self.contourThread = ContourDataThread(df_contour, self.contourDataPath,
                                                       roi_labels)
                del df_contour
                self.contourThread.output.connect(self.process_indicate)
                self.contourThread.finished.connect(self.save_data_indicate)
                self.contourThread.start()
            else:
                self.save_data_indicate()
                
    def save_data_indicate(self):
        self.statusBar.showMessage("Data saved successfully!")

    def process_indicate(self, status_text):
        self.statusBar.showMessage(status_text)

    def import_force_data(self): #import force data
        if self.msrListMode == False:
            self.forceData = ForceAnal()
        self.zeroForceData = ForceAnal() #initalise
        self.correctZeroForce.blockSignals(True)
        self.correctZeroForce.setChecked(False)
        self.correctZeroForce.blockSignals(False)        
        self.correctZeroForce.setEnabled(False)
        self.zeroForceFileNameLabel.setText("Select zero-force data from file menu")
        
        self.forceData.importData(self.msrListMode)
        if self.forceData.force_filepath != "":
            self.defl_vert1_raw = self.forceData.defl_vert1.copy() #copy of raw vert data
            self.defl_vert1_actual = self.forceData.defl_vert1.copy() #copy of raw vert data
            self.openZeroForceFile.setEnabled(True)
            self.fpsSpinBox.blockSignals(True)
            self.fpsSpinBox.setValue(self.forceData.fps)
            self.fpsSpinBox.blockSignals(False)
            self.frameTime = self.forceData.time_video #recalculated time array
            self.frameRate = self.forceData.fps
            self.forceFileNameLabel.setText("<b>Force data:</b>\n" + self.forceData.force_filepath)
##            self.videoEffect.model().item(6).setEnabled(True)
            self.init_plotconfig()
            self.forceData.dataClean()
            self.forceData.calcData()
            print(self.frameRate)

    def import_zero_force(self): #import zero force line
        self.zeroForceData = ForceAnal()
        self.correctZeroForce.setEnabled(True)
        self.zeroForceData.noiseSteps = ','.join(map(str,range(1,self.forceData.step_num + 1)))
        self.zeroForceData.importData(False)
        if self.zeroForceData.force_filepath != "":
            self.zeroForceFileNameLabel.setText("<b>Zero force data:</b>\n" +
                                                self.zeroForceData.force_filepath)
            self.zero_force_calc()

    def zero_force_calc(self): #calculate vertical force correction wrt zero
        zero_shift = self.zeroForceData.defl_vert1[0] - self.forceData.defl_vert1[0]
        self.zeroForceData.defl_vert1_corrected = [x-zero_shift for x in self.zeroForceData.defl_vert1]
##        self.defl_vert1_raw = self.forceData.defl_vert1.copy()
        self.defl_vert1_actual = [self.forceData.defl_vert1[0] +
                                 self.forceData.defl_vert1[i] -
                                 self.zeroForceData.defl_vert1_corrected[i] \
                                 for i in range(len(self.forceData.defl_vert1))]
        
    def zero_force_correct(self): #correct vertical force wrt zero
        if self.forceData.force_filepath != "":
            if self.correctZeroForce.isChecked() == True:
                self.zero_force_calc()
                self.forceData.defl_vert1 = self.defl_vert1_actual.copy()
            else:
                self.forceData.defl_vert1 = self.defl_vert1_raw.copy()
##        self.updatePlot(clean = False)

    def clear_data(self): #clear area data
        for k in self.roiDict.keys():
            if len(self.roiDict.keys()) > 1 and k == "Default":
                continue
            self.dataDict[k] = 6 * [np.zeros(int(self.frameCount), np.float64)] + \
                               [[[(0,0),(0,0),0,1]]*int(self.frameCount)]
        self.contour_data = [[], [], [], [], [], [], [], []]
        self.plotSequence()
        self.statusBar.showMessage("Data cleared!")

##    def summary_dialog(self): #dialog for summary combine
##        self.sumDialog = QDialog(self)
##        self.sumDialog.setWindowTitle("Choose grouping parameter")
##        self.sumDialog.resize(300, 300)
##        gridLayout = QGridLayout(self.sumDialog)
##        self.sumlistwidget = QListWidget(self.sumDialog)
##        okButton = QPushButton("Select", self.sumDialog)
##        okButton.clicked.connect(self.combine_summary_data)
##        okButton.setDefault(True)
##        
##        itemlist = ["Date", "Folder_Name", "Species", "Sex", "Leg", "Pad",
##                   "Weight", "Temperature", "Humidity", "Medium",
##                   "Substrate", "Data_OK", "Include_Data", "Label", "ROI Label"]
##        self.sumlistwidget.addItems(itemlist)
##
##        gridLayout.addWidget(self.sumlistwidget, 1, 0, 1, 1)
##        gridLayout.addWidget(okButton, 2, 0, 1, 1)
##        self.sumDialog.show()

    def summary_dialog_init(self): #initialise dialog for summary combine
        self.summary = None
        self.sumDialog = QDialog(self)
        self.sumDialog.setWindowTitle("Configure Summary Plots")
##        self.sumDialog.resize(300, 300)
        
        plotTitleLabel = QLabel("<b>Plot Title</b>", self.sumDialog)
        plotLabel = QLabel("<b>Plot Number</b>", self.sumDialog)
        xLabel = QLabel("<b>X Variable</b>", self.sumDialog)
        yLabel = QLabel("<b>Y Variable</b>", self.sumDialog)
        colorbarLabel = QLabel("<b>Colorbar Variable</b>", self.sumDialog)
        groupVarLabel = QLabel("<b>Group By</b>", self.sumDialog)
        formatLabel = QLabel("<b>Format</b>", self.sumDialog)
        fitLabel = QLabel("<b>Fit</b>", self.sumDialog)
        orderLabel = QLabel("<b>Order</b>", self.sumDialog)

        self.summaryDict = {'x var':[None, None, None, None],
                            'y var':[None, None, None, None],
                            'cbar var':[None, None, None, None],
                            'plot num':[None, None, None, None],
                            'fit': [None, None, None, None],
                            'order': [None, None, None, None],
                            'title':[None], 'format':[None]} #initialize

        plotTitle =  QLineEdit(self.sumDialog)
        plotTitle.textChanged.connect(lambda: self.update_summary_dict('title',
                                                                       plotTitle.text(),0))
        plotTitle.setText('Adhesion vs Area')
        
        plotFormat = QComboBox(self.sumDialog) #plot save format
        plotFormat.addItems(['jpg', 'svg', 'pdf', 'png', 'tif', 'tiff'])
        plotFormat.currentIndexChanged.connect(lambda:
                                               self.update_summary_dict('format',
                                                                        plotFormat.currentText(), 0))
        self.update_summary_dict('format', plotFormat.currentText(), 0)
        
        grouplist = ["Date", "Folder_Name", "Species", "Sex", "Leg", "Pad",
                   "Weight", "Temperature", "Humidity", "Medium",
                   "Substrate", "Data_OK", "Include_Data", "Label", "ROI Label",
                     "Measurement_Number", "Contact_Time", "Detachment Speed",
                   "Attachment Speed", "Sliding Speed", "Sliding_Step"]
        grouplist.sort()

        groupVar = QComboBox(self.sumDialog)
        groupVar.addItems(grouplist)
        # ind = groupVar.findText("ROI Label")
        groupVar.setCurrentIndex(groupVar.findText("ROI Label"))
##        groupVar.setEnabled(False)

        combine = QCheckBox('Combine', self.sumDialog)
        okButton = QPushButton("Select summary file..", self.sumDialog)

        combine.stateChanged.connect(lambda: self.combine_toggled(groupVar, combine, okButton))       
        okButton.clicked.connect(lambda: self.combine_summary_data(combine.isChecked(),
                                                                   groupVar.currentText()))
##        okButton.setDefault(True)
        resetButton = QPushButton("Reset", self.sumDialog)
        resetButton.clicked.connect(self.reset_summary)

        gridLayout = QGridLayout(self.sumDialog)
        
        gridLayout.addWidget(plotTitleLabel, 0, 0, 1, 1)
        gridLayout.addWidget(plotTitle, 0, 1, 1, 1)
        gridLayout.addWidget(plotLabel, 1, 0, 1, 1)
        gridLayout.addWidget(xLabel, 1, 1, 1, 1)
        gridLayout.addWidget(yLabel, 1, 2, 1, 1)
        gridLayout.addWidget(colorbarLabel, 1, 3, 1, 1)
        gridLayout.addWidget(fitLabel, 1, 4, 1, 1, alignment = Qt.AlignCenter)
        gridLayout.addWidget(orderLabel, 1, 5, 1, 1)
        self.summary_layout_make(1, 'Pulloff_Area', 'Adhesion_Force',
                                 'Detachment Speed', gridLayout, 2)
        self.summary_layout_make(2, 'Pulloff_Area', 'Adhesion_Force',
                                 'Adhesion_Preload', gridLayout, 3)
        self.summary_layout_make(3, 'Pulloff_Area', 'Adhesion_Force',
                                 'Contact_Time', gridLayout, 4)
        self.summary_layout_make(4, 'Pulloff_Area', 'Adhesion_Force',
                                 'Sliding_Step', gridLayout, 5)
        gridLayout.addWidget(combine, 0, 2, 1, 1)
        gridLayout.addWidget(groupVarLabel, 0, 2, 1, 1, alignment = Qt.AlignRight)
        gridLayout.addWidget(groupVar, 0, 3, 1, 1)
        gridLayout.addWidget(formatLabel, 0, 4, 1, 1, alignment = Qt.AlignRight)
        gridLayout.addWidget(plotFormat, 0, 5, 1, 1)
        gridLayout.addWidget(okButton, 6, 3, 1, 1)
        gridLayout.addWidget(resetButton, 6, 1, 1, 1)
        
##        self.sumDialog.show()

    def summary_layout_make(self, plotnum, x_init, y_init, cb_init,
                            layout, vpos):
        
        varlist = ["Adhesion_Force", "Adhesion_Preload", "Friction_Force",
                   "Friction_Preload", "Max_Area", "Pulloff_Area",
                   "Friction_Area", "ROI_Max_Area", "ROI_Pulloff_Area",
                   "Max_Length", "Pulloff_Length", "ROI_Max_Length",
                   "ROI_Pulloff_Length", "Pulloff_Contact_Number",
                   "Residue_Area", "Pulloff_Median_Eccentricity", "ROI Label",
                   "Measurement_Number", "Contact_Time", "Detachment Speed",
                   "Attachment Speed", "Sliding Speed", "Sliding_Step", "Slope",
                   "Adhesion_Stress", "Friction_Stress", "Beam_Spring_Constant",
                   "Initial_Deformation","Pulloff_Deformation","Adhesion_Energy",
                   "Max_Bounding_Area", "Max_Bounding_Perimeter",
                   "Max_Bounding_Length", "Max_Bounding_Width", 
                   "Adhesion_Energy_per_Area", "Date_of_Experiment"]
        varlist.sort()
        plotLabel = QLabel(str(plotnum), self.sumDialog)
        self.update_summary_dict('plot num', plotnum, plotnum-1)

        xVar = QComboBox(self.sumDialog) #x variable
        xVar.addItems(varlist)
        xVar.currentIndexChanged.connect(lambda: self.update_summary_dict('x var',
                                                                          xVar.currentText(),
                                                                          plotnum-1))
        xVar.setCurrentIndex(xVar.findText(x_init))
        self.update_summary_dict('x var', xVar.currentText(), plotnum-1)

        yVar = QComboBox(self.sumDialog) #y variable
        yVar.addItems(varlist)
        yVar.currentIndexChanged.connect(lambda: self.update_summary_dict('y var',
                                                                          yVar.currentText(),
                                                                          plotnum-1))
        yVar.setCurrentIndex(yVar.findText(y_init))
        self.update_summary_dict('y var', yVar.currentText(), plotnum-1)
        
        colorbarVar = QComboBox(self.sumDialog) #colorbar variable
        colorbarVar.addItems(varlist)
        colorbarVar.currentIndexChanged.connect(lambda: self.update_summary_dict('cbar var',
                                                                                 colorbarVar.currentText(),
                                                                                 plotnum-1))
        colorbarVar.setCurrentIndex(colorbarVar.findText(cb_init))
        self.update_summary_dict('cbar var', colorbarVar.currentText(), plotnum-1)

        polyfit = QCheckBox(self.sumDialog) #polynomial fit
        polyfit.stateChanged.connect(lambda: self.update_summary_dict('fit',
                                                                      polyfit.isChecked(),
                                                                      plotnum-1))
        self.update_summary_dict('fit', polyfit.isChecked(), plotnum-1)

        polyorder = QSpinBox(self.sumDialog) #polynomial order
        polyorder.valueChanged.connect(lambda: self.update_summary_dict('order',
                                                                      polyorder.value(),
                                                                      plotnum-1))
        polyorder.setRange(1, 10)
        polyorder.setValue(1)
        self.update_summary_dict('order', polyorder.value(), plotnum-1)
        
        layout.addWidget(plotLabel, vpos, 0, 1, 1)
        layout.addWidget(xVar, vpos, 1, 1, 1)
        layout.addWidget(yVar, vpos, 2, 1, 1)
        layout.addWidget(colorbarVar, vpos, 3, 1, 1)
        layout.addWidget(polyfit, vpos, 4, 1, 1, alignment = Qt.AlignCenter)
        layout.addWidget(polyorder, vpos, 5, 1, 1)

    def combine_toggled(self, groupVar, combine, okButton):
##        groupVar.setEnabled(combine.isChecked())
        oktext = "Select summary file.." if combine.isChecked() == False \
                 else "Select experiment list.."
        okButton.setText(oktext)

    def update_summary_dict(self, key, value, plotnum):
        self.summaryDict[key][plotnum] = value
        
    def combine_summary_data(self, combine, legend_parameter): #combine summary plots
##        self.sumDialog.reject()
##        legend_parameter = self.sumlistwidget.currentItem().text()
        self.reset_summary()
        
        if combine == True:
            self.summary = SummaryAnal()
            self.summary.combineSummary(self.summaryDict, legend_parameter)
            if self.summary.list_filepath != "":
                self.comb = True
                self.statusBar.showMessage("Summary Data combined!")
            else:
                self.statusBar.showMessage("No file selected")
                self.comb = False
        else:
            self.comb = False

        self.show_summary_plots(legend_parameter)
        
    def export_summary_plots(self): #export summary plots
        if self.comb == False and self.summary == None:
            self.reset_summary()
            self.summary = SummaryAnal()
            self.summary.importSummary()
##            if len(self.summary.roi_label_unique) == 2:
##                self.summary.roi_label_unique.remove("All")
##        self.summary.eq_count = [1,1,1,1]
            self.summary.plotSummary(self.summaryDict,
                                     self.summary.df_forcedata,
                                     self.summary.df_forcedata)
        #save summary plots in separate thread
        saveSummPlotThread = SummPlotThread(self.summary,
                                            self.summaryDict['format'][0])
        saveSummPlotThread.output.connect(self.process_indicate)
        saveSummPlotThread.finished.connect(self.save_plot_indicate)
        saveSummPlotThread.start()

    def save_plot_indicate(self):
        self.statusBar.showMessage("Summary Plots saved!")

    def show_summary_plots(self, group = "ROI Label"): #show summary plots
        if self.comb == False and self.summary == None:
            self.reset_summary()
            self.summary = SummaryAnal()
            self.summary.importSummary()
##            if len(self.summary.roi_label_unique) == 2:
##                self.summary.roi_label_unique.remove("All")
##        self.summary.eq_count = [1,1,1,1]
            self.summary.plotSummary(self.summaryDict,
                                     self.summary.df_forcedata,
                                     self.summary.df_forcedata,
                                     group)
        self.summary.showSummaryPlot()

    def reset_summary(self): #reset self.comb to False
        self.comb = False
        self.summary = None
##        self.summary_dialog_init()
        plt.clf()
        plt.cla()
        plt.close()
        gc.collect()
        self.statusBar.showMessage("Reset!")
           
    def closeEvent(self, event): #close application
        msg = QMessageBox()
        msg.setStyleSheet("QLabel{min-width:500 px; font-size: 24px;} QPushButton{ width:500px; font-size: 24px; }");
        choice = msg.question(self, 'Closing...',
                                      "Really Quit?",
                                      msg.Yes | msg.No)        
        if choice == msg.Yes:
            if type(event) is not bool:
                event.accept()
            self.frameAction = "Stop" #exit play loop
            self.reset_summary()
            cv2.destroyAllWindows()
            QApplication.exit()
        else:
            if type(event) is not bool:
                event.ignore() 

# %% Configure ROIs window for analysis
class ConfigROIWindow(QWidget): 
    def __init__(self):
        super().__init__()
        self.setGeometry(720, 140, 100, 100)
        self.setWindowTitle("Configure ROI")
        self.layout = QGridLayout()
        self.roiDict = {}
        self.home()

    def home(self):
        self.roiNum = QSpinBox(self) #no. of ROI
        self.roiNum.setValue(1)
        self.roiNum.setSingleStep(1)
        self.roiNum.setRange(1, 100)
        self.roiNum.valueChanged.connect(self.num_change)
        self.label1 = QLabel("ROI Number", self)

        self.roiLabel =  QLineEdit(self) #Name of ROI
##        self.roiLabel.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.roiLabel.setText("")
        self.roiLabel.textChanged.connect(self.label_change)
        self.label2 = QLabel("ROI Name", self)

        self.roiDrawBtn = QPushButton("Draw ROI", self) #Draw ROI
        self.roiDrawBtn.setEnabled(False)
##        self.roiDrawBtn.clicked.connect(self.roi_draw)

        self.roiDef = QLabel("ROI Definition:", self) #display definitions

        self.delBtn = QPushButton("Delete", self) #Delete current definition
        self.delBtn.clicked.connect(self.del_roi)

        self.okBtn = QPushButton("OK", self) #Close window
        
        self.layout.addWidget(self.roiNum, 1, 0, 1, 1)
        self.layout.addWidget(self.label1, 0, 0, 1, 1)
        self.layout.addWidget(self.roiLabel, 1, 1, 1, 1)
        self.layout.addWidget(self.label2, 0, 1, 1, 1)
        self.layout.addWidget(self.roiDrawBtn, 2, 0, 1, 1)
        self.layout.addWidget(self.delBtn, 2, 1, 1, 1)
        self.layout.addWidget(self.roiDef, 0, 2, 3, 1)
        self.layout.addWidget(self.okBtn, 3, 0, 1, 2)

        self.setLayout(self.layout)

    def num_change(self):
        key = self.roiNum.value()
        self.roiLabel.blockSignals(True)
        if key in self.roiDict.keys():
            self.roiLabel.setText(self.roiDict[key])
            self.roiDrawBtn.setEnabled(True)
        else:
            self.roiLabel.setText("")
            self.roiDrawBtn.setEnabled(False)
        self.roiLabel.blockSignals(False)

    def label_change(self):
        if self.roiLabel.text() in ["Dict", "dict", "All", "all"]: #banned keywords
            self.roiLabel.blockSignals(True)
            self.roiLabel.setText("")
            self.roiLabel.blockSignals(False)
            self.roiDict[self.roiNum.value()] = self.roiLabel.text()
            self.update_def()
            self.roiDrawBtn.setEnabled(False)
        else:
            self.roiDict[self.roiNum.value()] = self.roiLabel.text()
            self.update_def()
            self.roiDrawBtn.setEnabled(True)

    def del_roi(self): #delete current definition
        del self.roiDict[self.roiNum.value()]
        self.roiNum.setValue(self.roiNum.value()-1)
        self.update_def()

    def update_def(self): #update definition label
        defString = "ROI Definition:"
        for k in self.roiDict.keys():
            defString += "\nROI " + str(k) + ":\t" + self.roiDict[k]

        self.roiDef.setText(defString)
        
    def roi_draw(self):
        pass

    def showWindow(self):
        self.show()

# %% Configure file paths window
class ConfigPathWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.setGeometry(100, 100, 500, 500)
        self.setWindowTitle("Configure Paths")
        self.home()

    def home(self):
        self.browseBtn1 = QPushButton("Browse..", self) #data
        self.browseBtn1.clicked.connect(lambda: self.browse_folder('data'))
        self.dataPath =  QTextEdit(self) #data
        self.dataPath.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

        self.browseBtn2 = QPushButton("Browse..", self) #plots
        self.browseBtn2.clicked.connect(lambda: self.browse_folder('plot'))        
        self.plotPath =  QTextEdit(self) #plots
        self.plotPath.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

        self.browseBtn3 = QPushButton("Browse..", self) #recording
        self.browseBtn3.clicked.connect(lambda: self.browse_folder('recording'))
        self.recordingPath =  QTextEdit(self) #recording
        self.recordingPath.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

        self.browseBtn4 = QPushButton("Browse..", self) #summary
        self.browseBtn4.clicked.connect(lambda: self.browse_folder('summary'))
        self.summaryPath =  QTextEdit(self) #summary
        self.summaryPath.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

        self.browseBtn5 = QPushButton("Browse..", self) #contours
        self.browseBtn5.clicked.connect(lambda: self.browse_folder('contours'))
        self.contourDataPath =  QTextEdit(self) #contours
        self.contourDataPath.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        
        self.okBtn = QPushButton("OK", self) #Close window

        self.dataGroupBox = self.createGroup("Data", self.dataPath,
                                             self.browseBtn1)
        self.plotGroupBox = self.createGroup("Plot", self.plotPath,
                                             self.browseBtn2)
        self.recordingGroupBox = self.createGroup("Recording", self.recordingPath,
                                             self.browseBtn3)
        self.summaryGroupBox = self.createGroup("Summary", self.summaryPath,
                                             self.browseBtn4)
        self.contourGroupBox = self.createGroup("Contour Data", self.contourDataPath,
                                             self.browseBtn5)

        self.layout = QVBoxLayout()
        self.setLayout(self.layout)
        self.layout.addWidget(self.dataGroupBox)
        self.layout.addWidget(self.plotGroupBox)
        self.layout.addWidget(self.recordingGroupBox)
        self.layout.addWidget(self.summaryGroupBox)
        self.layout.addWidget(self.contourGroupBox)
        self.layout.addWidget(self.okBtn)

    def createGroup(self, title, path, button): #create groupbox
        groupBox = QGroupBox(title)
        groupBox.setStyleSheet("QGroupBox { font-weight: bold; } ")
        groupBox.setCheckable(True)
        groupBox.setChecked(True)
        hbox = QHBoxLayout(self)
        groupBox.setLayout(hbox)
        hbox.addWidget(path)
        hbox.addWidget(button)
        return groupBox
    
    def browse_folder(self, filetype):
        name, _ = QFileDialog.getSaveFileName(self, "Save " + filetype + " file")
        if filetype == 'data':
            self.dataPath.setText(name + '.txt')
        elif filetype == 'plot':
            self.plotPath.setText(name + '.svg')
        elif filetype == 'recording':
            self.recordingPath.setText(name + '.avi')
        elif filetype == 'summary':
            self.summaryPath.setText(name + '.txt')
        elif filetype == 'contours':
            self.contourDataPath.setText(name + '.xlsx')

    def showWindow(self):
        self.show()

# %% Recording Window
class ConfigRecWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.setGeometry(100, 100, 550, 250)
        self.setWindowTitle("Configure Record")
        self.layout = QGridLayout()
        self.cap2 = None
        self.home()

    def home(self):
        self.saveBtn = QPushButton("Browse..", self) #Save Video
        self.saveBtn.clicked.connect(self.save_dialog)

        self.openBtn = QPushButton("Open...", self) #Open 2nd Video
        self.openBtn.clicked.connect(self.open_video)

        self.videoTextbox =  QTextEdit(self) #2nd video filename diplayed
        self.videoTextbox.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

        self.fourRec = QCheckBox('4 panel recording', self) #record 4 views

        self.textbox =  QTextEdit(self) #filename diplayed
##        selt.textbox.wordWrapMode()
        self.textbox.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        
        self.codec = QComboBox(self) #length unit
        self.codec.addItem("MJPG")
        self.codec.addItem("DIVX")
        self.codec.addItem("FFV1")
        self.codec.setCurrentIndex(1)
        self.codecLabel = QLabel("Codec:", self)

        self.fps = QSpinBox(self) #playback fps of recorded file
        self.fps.setValue(1)
        self.fps.setSingleStep(1)
        self.fps.setRange(1, 100)
        self.fpsLabel = QLabel("Frames per second:", self)
        
        self.okBtn = QPushButton("OK", self) #Close window

        fileGroupBox = QGroupBox("Save video in...")
        codecGroupBox = QGroupBox("Video settings")
        videoGroupBox = QGroupBox("Select 2nd video")

        fileGroupBox.setStyleSheet("QGroupBox { font-weight: bold; } ")
        codecGroupBox.setStyleSheet("QGroupBox { font-weight: bold; } ")
        videoGroupBox.setStyleSheet("QGroupBox { font-weight: bold; } ")
        
        self.layout.addWidget(fileGroupBox, 0, 0)
        self.layout.addWidget(codecGroupBox, 2, 0)
        self.layout.addWidget(videoGroupBox, 0, 1)
        self.layout.addWidget(self.okBtn, 2, 1)

        self.setLayout(self.layout)
        
        fileVbox = QGridLayout()
        fileGroupBox.setLayout(fileVbox)
        fileVbox.addWidget(self.saveBtn, 0, 0, 1, 2)
        fileVbox.addWidget(self.textbox, 1, 0, 2, 2)

        videoVbox = QGridLayout()
        videoGroupBox.setLayout(videoVbox)
        videoVbox.addWidget(self.openBtn, 0, 0, 1, 1)
        videoVbox.addWidget(self.fourRec, 0, 1, 1, 1)
        videoVbox.addWidget(self.videoTextbox, 1, 0, 2, 2)

        codecVbox = QGridLayout()
        codecGroupBox.setLayout(codecVbox)
        codecVbox.addWidget(self.codecLabel, 0, 0, 1, 1)
        codecVbox.addWidget(self.codec, 0, 1, 1, 1)
        codecVbox.addWidget(self.fpsLabel, 1, 0, 1, 1)
        codecVbox.addWidget(self.fps, 1, 1, 1, 1)

    def save_dialog(self):
        name, _ = QFileDialog.getSaveFileName(self, "Save File")
        filename = time.strftime(name + "-%Y%m%d%H%M%S" + ".avi")
        self.textbox.setText(filename)

    def open_video(self):
        videofile, _ = QFileDialog.getOpenFileName(self, "Open Video")
        self.videoTextbox.setText(videofile)

    def showWindow(self, filepath):
##        self.filename = filepath + "-recording" + ".avi"
        self.textbox.setText(filepath)
        self.show()

# %% Configure Plot Window
class ConfigPlotWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.setGeometry(100, 580, 1000, 200)
        self.setWindowTitle("Configure Plot")
        self.layout = QGridLayout()
        self.home()
        self.rangeDict = {"Default" : [[0,1],[0,100],[0,100],
                                       [0,100],[0,100],[0,1]]}

    def home(self):

        self.showContactArea = QCheckBox('contact area', self) #contact area
        self.showContactArea.setChecked(True)

        self.showROIArea = QCheckBox('ROI area', self) #roi area
        self.showContactLength = QCheckBox('contact length', self) #contact length
        self.showROILength = QCheckBox('ROI length', self) #roi length
        self.showContactNumber = QCheckBox('contact number', self) #contact number
        self.showEcc = QCheckBox('eccentricity', self) #median eccentricity

        self.showLateralForce = QCheckBox('lateral force', self) #lateral force
        self.showZPiezo = QCheckBox('vertical piezo', self) #z piezo
        self.showXPiezo = QCheckBox('lateral piezo', self) #x piezo
        self.showAdhesion = QCheckBox('adhesion calculation', self) #adhesion/preload calc line
        self.showFriction = QCheckBox('friction calculation', self) #friction calc lines
        self.showStress = QCheckBox('stress', self) #stress
        self.showDeformation = QCheckBox('deformation', self) #deformation

        self.xAxisLabel = QLabel("<b>X Axis:</b>", self)
        self.xAxisParam = QComboBox(self) #x axis parameter
        self.xAxisParam.addItem("Time (s)")
        self.xAxisParam.addItem("Vertical Position (μm)")
        self.xAxisParam.addItem("Lateral Position (μm)")
        self.xAxisParam.addItem("Deformation (μm)")
        
        self.roiChoice = QComboBox(self) #choose ROI
        self.roiChoice.addItem("Default")
        self.roiChoice.setCurrentIndex(0)
        self.roiChoice.currentIndexChanged.connect(self.update_range)
        
        self.startLabel = QLabel("Start (%):", self)
        self.endLabel = QLabel("End (%):", self)

        self.zeroLabel = QLabel("Zero Range", self)
        self.adhLabel = QLabel("Adhesion Range", self)
        self.prl1Label = QLabel("Preload Range", self)
        
        self.zeroRange1 = QDoubleSpinBox(self) #vertical force zero range start
        self.zeroRange1.setValue(0)
        self.zeroRange1.setSingleStep(1)
        self.zeroRange1.setRange(0, 100)
        self.zeroRange1.valueChanged.connect(self.update_dict)
        
        self.zeroRange2 = QDoubleSpinBox(self) #vertical force zero range end
        self.zeroRange2.setValue(1)
        self.zeroRange2.setSingleStep(1)
        self.zeroRange2.setRange(0, 100)
        self.zeroRange2.valueChanged.connect(self.update_dict)

        self.adhRange1 = QDoubleSpinBox(self) #adhesion peak range start
        self.adhRange1.setValue(0)
        self.adhRange1.setSingleStep(1)
        self.adhRange1.setRange(0, 100)
        self.adhRange1.valueChanged.connect(self.update_dict)

        self.adhRange2 = QDoubleSpinBox(self) #adhesion peak range start
        self.adhRange2.setValue(100)
        self.adhRange2.setSingleStep(1)
        self.adhRange2.setRange(0, 100)
        self.adhRange2.valueChanged.connect(self.update_dict)

        self.prl1Range1 = QDoubleSpinBox(self) #preload peak range start
        self.prl1Range1.setValue(0)
        self.prl1Range1.setSingleStep(1)
        self.prl1Range1.setRange(0, 100)
        self.prl1Range1.valueChanged.connect(self.update_dict)

        self.prl1Range2 = QDoubleSpinBox(self) #preload peak range start
        self.prl1Range2.setValue(100)
        self.prl1Range2.setSingleStep(1)
        self.prl1Range2.setRange(0, 100)
        self.prl1Range2.valueChanged.connect(self.update_dict)

        self.zero2Range1 = QDoubleSpinBox(self) #lateral force zero range start
        self.zero2Range1.setValue(0)
        self.zero2Range1.setSingleStep(1)
        self.zero2Range1.setRange(0, 100)
        self.zero2Range1.valueChanged.connect(self.update_dict)

        self.zero2Range2 = QDoubleSpinBox(self) #lateral force zero range end
        self.zero2Range2.setValue(1)
        self.zero2Range2.setSingleStep(1)
        self.zero2Range2.setRange(0, 100)
        self.zero2Range2.valueChanged.connect(self.update_dict)

        self.filterLatF = QCheckBox('Filter cyan curve', self) #filter
        
        self.filter_wind = QSpinBox(self) #filter window
        self.filter_wind.setValue(43)
        self.filter_wind.setSingleStep(20)
        self.filter_wind.setRange(3, 10001)
        self.filter_wind.valueChanged.connect(self.filter_change)
        self.windLabel = QLabel("Window Length:", self)
        
        self.filter_poly = QSpinBox(self) #filter polynom
        self.filter_poly.setValue(2)
        self.filter_poly.setSingleStep(1)
        self.filter_poly.setRange(1, 20000)
        self.polyLabel = QLabel("Polynomial Order:", self)

        self.startLabel2 = QLabel("Start (%):", self)
        self.endLabel2 = QLabel("End (%):", self)

        self.frLabel = QLabel("Friction Range", self)
        self.prl2Label = QLabel("Preload Range", self)
        self.zero2Label = QLabel("Zero Range", self)
        
        self.eqLabel = QLabel("Lateral Calib. Equation (μN):", self)        
        self.latCalibEq = QLineEdit(self) #lateral force calib equation
        self.latCalibEq.setText("29181.73*x")

        self.noiseStepsLabel = QLabel("Noisy Steps:", self)        
        self.noiseSteps = QLineEdit(self) #remove first data point from steps
        self.noiseSteps.setText("")

        self.legendPosLabel = QLabel("Legend:", self) #legend position        
        self.legendPos = QLineEdit(self)
        self.legendPos.setText("upper right")

        self.startFullLabel = QLabel("Start (%):", self)
        self.endFullLabel = QLabel("End (%):", self)

        self.startFull = QDoubleSpinBox(self) #plot range start
        self.startFull.setValue(0)
        self.startFull.setSingleStep(1)
        self.startFull.setRange(0, 100)

        self.endFull = QDoubleSpinBox(self) #plot range end
        self.endFull.setValue(100)
        self.endFull.setSingleStep(1)
        self.endFull.setRange(0, 100)
        
        self.invertLatForce = QCheckBox('Invert Lateral Force', self) #invert

        self.applyCrossTalk = QCheckBox('Apply Cross Talk', self) #cross talk flag
        self.zeroShift = QCheckBox('Shift to Zero', self) #force curve shift to zero

        self.vertCrossTalk = QDoubleSpinBox(self) #vertical cross talk slope
        self.vertCrossTalk.setValue(0)
        self.vertCrossTalk.setSingleStep(0.1)
        self.vertCrossTalk.setDecimals(4)
        self.vertCrossTalk.setRange(-1000, 1000)
        self.vertCTlabel = QLabel("Cross Talk (μN/μN):", self)

        self.latCrossTalk = QDoubleSpinBox(self) #lateral cross talk slope
        self.latCrossTalk.setValue(0)
        self.latCrossTalk.setSingleStep(0.1)
        self.latCrossTalk.setDecimals(4)
        self.latCrossTalk.setRange(-1000, 1000)
        self.latCTlabel = QLabel("Cross Talk (μN/μN):", self)

        self.frictionRange1 = QDoubleSpinBox(self) #friction range start
        self.frictionRange1.setValue(0)
        self.frictionRange1.setSingleStep(1)
        self.frictionRange1.setRange(0, 100)
        self.frictionRange1.valueChanged.connect(self.update_dict)
        
        self.frictionRange2 = QDoubleSpinBox(self) #friction range end
        self.frictionRange2.setValue(100)
        self.frictionRange2.setSingleStep(1)
        self.frictionRange2.setRange(0, 100)
        self.frictionRange2.valueChanged.connect(self.update_dict)

        self.prl2Range1 = QDoubleSpinBox(self) #friction preload peak range start
        self.prl2Range1.setValue(0)
        self.prl2Range1.setSingleStep(1)
        self.prl2Range1.setRange(0, 100)
        self.prl2Range1.valueChanged.connect(self.update_dict)

        self.prl2Range2 = QDoubleSpinBox(self) #friction preload peak range start
        self.prl2Range2.setValue(100)
        self.prl2Range2.setSingleStep(1)
        self.prl2Range2.setRange(0, 100)
        self.prl2Range2.valueChanged.connect(self.update_dict)

        self.startFitLabel = QLabel("Start (%):", self)
        self.endFitLabel = QLabel("End (%):", self)

        self.fitStart = QDoubleSpinBox(self) #fitting range start
        self.fitStart.setValue(0)
        self.fitStart.setSingleStep(1)
        self.fitStart.setRange(0, 100)

        self.fitStop = QDoubleSpinBox(self) #fitting range end
        self.fitStop.setValue(100)
        self.fitStop.setSingleStep(1)
        self.fitStop.setRange(0, 100)

        self.xFitLabel = QLabel("X Parameter:", self)
        self.yFitLabel = QLabel("Y Parameter:", self)

        self.xFit = QComboBox(self) #x param
        self.xFit.addItems(['Deformation (μm)',
                            'Vertical Position (μm)',
                            'Lateral Position (μm)',
                            'Time (s)'])
        self.xFit.setCurrentIndex(0)

        self.yFit = QComboBox(self) #x param
        self.yFit.addItems(['Vertical Force (μN)', 'Lateral Force (μN)'])
        self.yFit.setCurrentIndex(0)

        self.fitPosLabel = QLabel("Fit Position\n(x,y):", self) #fit eq. position        
        self.fitPos = QLineEdit(self)
        self.fitPos.setText('0.5,0.5')

        self.showFitEq = QCheckBox('Show Slope', self) #display equation on plot
        
        self.kBeamLabel = QLabel("Beam Spring Constant (μN/μm):", self) #beam dpring constant       
        self.kBeam = QLineEdit(self)
        self.kBeam.setText('30,1')

        self.deformStartLabel = QLabel("Deformation Start:", self) #contact start tolerance auto detect
        self.deformStart = QSpinBox(self) 
        self.deformStart.setValue(100)
        self.deformStart.setSingleStep(1)
        self.deformStart.setRange(0, 10000)
        
        self.okBtn = QPushButton("OK", self) #Close window

        self.updateBtn = QPushButton("Update", self) #Update
        
        self.zeroGroupBox = QGroupBox("Configure Vertical Force")
        filterGroupBox = QGroupBox("Configure Plot")
        flagGroupBox = QGroupBox("Show")
        self.latCalibGroupBox = QGroupBox("Configure Lateral Force")
        self.fittingGroupBox = QGroupBox("Fit Data")
        buttonGroupBox = QGroupBox()

        self.zeroGroupBox.setStyleSheet("QGroupBox { font-weight: bold; } ")
        filterGroupBox.setStyleSheet("QGroupBox { font-weight: bold; } ")
        flagGroupBox.setStyleSheet("QGroupBox { font-weight: bold; } ")
        self.latCalibGroupBox.setStyleSheet("QGroupBox { font-weight: bold; } ")
        self.fittingGroupBox.setStyleSheet("QGroupBox { font-weight: bold; } ")
        self.fittingGroupBox.setCheckable(True)
        self.fittingGroupBox.setChecked(False)

        self.layout.addWidget(self.roiChoice, 0, 0, 1, 2)
        self.layout.addWidget(self.zeroGroupBox, 1, 0)
        self.layout.addWidget(filterGroupBox, 2, 1)
        self.layout.addWidget(flagGroupBox, 2, 0)
        self.layout.addWidget(self.latCalibGroupBox, 1, 1)
        self.layout.addWidget(self.fittingGroupBox, 3, 0)
        self.layout.addWidget(buttonGroupBox, 3, 1)

        self.setLayout(self.layout)

        buttonVbox = QGridLayout()
        buttonGroupBox.setLayout(buttonVbox)
        buttonVbox.addWidget(self.updateBtn, 0, 0)
        buttonVbox.addWidget(self.okBtn, 0, 1)
        
        zeroVbox = QGridLayout()
        self.zeroGroupBox.setLayout(zeroVbox)
        zeroVbox.addWidget(self.zeroLabel, 0, 1, 1, 1)
        zeroVbox.addWidget(self.adhLabel, 0, 2, 1, 1)
        zeroVbox.addWidget(self.prl1Label, 0, 3, 1, 1)
        zeroVbox.addWidget(self.startLabel, 1, 0, 1, 1)
        zeroVbox.addWidget(self.endLabel, 2, 0, 1, 1)
        zeroVbox.addWidget(self.zeroRange1, 1, 1, 1, 1)
        zeroVbox.addWidget(self.zeroRange2, 2, 1, 1, 1)
        zeroVbox.addWidget(self.adhRange1, 1, 2, 1, 1)
        zeroVbox.addWidget(self.adhRange2, 2, 2, 1, 1)
        zeroVbox.addWidget(self.prl1Range1, 1, 3, 1, 1)
        zeroVbox.addWidget(self.prl1Range2, 2, 3, 1, 1)
        zeroVbox.addWidget(self.vertCTlabel, 3, 0, 1, 1)
        zeroVbox.addWidget(self.vertCrossTalk, 3, 1, 1, 1)

        filterVbox = QGridLayout()
        filterGroupBox.setLayout(filterVbox)
        filterVbox.addWidget(self.filterLatF, 1, 0, 1, 2)
        filterVbox.addWidget(self.windLabel, 2, 0, 1, 1)
        filterVbox.addWidget(self.filter_wind, 2, 1, 1, 1)
        filterVbox.addWidget(self.polyLabel, 3, 0, 1, 1)
        filterVbox.addWidget(self.filter_poly, 3, 1, 1, 1)
        filterVbox.addWidget(self.eqLabel, 2, 2, 1, 2)
        filterVbox.addWidget(self.latCalibEq, 3, 2, 1, 2)
        filterVbox.addWidget(self.invertLatForce, 0, 2, 1, 2)
        filterVbox.addWidget(self.zeroShift, 0, 0, 1, 1)
        filterVbox.addWidget(self.applyCrossTalk, 1, 2, 1, 2)
        filterVbox.addWidget(self.xAxisLabel, 0, 3, 1, 1)
        filterVbox.addWidget(self.xAxisParam, 1, 3, 1, 1)
        filterVbox.addWidget(self.noiseStepsLabel, 4, 2, 1, 1)
        filterVbox.addWidget(self.noiseSteps, 5, 2, 1, 1)
        filterVbox.addWidget(self.legendPosLabel, 4, 3, 1, 1)
        filterVbox.addWidget(self.legendPos, 5, 3, 1, 1)
        filterVbox.addWidget(self.startFullLabel, 4, 0, 1, 1)
        filterVbox.addWidget(self.endFullLabel, 5, 0, 1, 1)
        filterVbox.addWidget(self.startFull, 4, 1, 1, 1)
        filterVbox.addWidget(self.endFull, 5, 1, 1, 1)
        filterVbox.addWidget(self.kBeamLabel, 6, 2, 1, 1)
        filterVbox.addWidget(self.kBeam, 6, 3, 1, 1)
        filterVbox.addWidget(self.deformStartLabel, 6, 0, 1, 1)
        filterVbox.addWidget(self.deformStart, 6, 1, 1, 1)

        flagVbox = QGridLayout()
        flagGroupBox.setLayout(flagVbox)
        flagVbox.addWidget(self.showContactArea, 0, 0)
        flagVbox.addWidget(self.showROIArea, 0, 1)
        flagVbox.addWidget(self.showZPiezo, 0, 2)
        flagVbox.addWidget(self.showXPiezo, 1, 0)
        flagVbox.addWidget(self.showAdhesion, 1, 1)
        flagVbox.addWidget(self.showFriction, 1, 2)
        flagVbox.addWidget(self.showLateralForce, 2, 0)
        flagVbox.addWidget(self.showContactLength, 2, 1)
        flagVbox.addWidget(self.showROILength, 2, 2)
        flagVbox.addWidget(self.showContactNumber, 3, 0)
        flagVbox.addWidget(self.showEcc, 3, 1)
        flagVbox.addWidget(self.showStress, 3, 2)
        flagVbox.addWidget(self.showDeformation, 4, 0)

        lastCalibVbox = QGridLayout()
        self.latCalibGroupBox.setLayout(lastCalibVbox)
        lastCalibVbox.addWidget(self.frLabel, 0, 1, 1, 1)
        lastCalibVbox.addWidget(self.prl2Label, 0, 2, 1, 1)
        lastCalibVbox.addWidget(self.zero2Label, 0, 3, 1, 1)
        lastCalibVbox.addWidget(self.startLabel2, 1, 0, 1, 1)
        lastCalibVbox.addWidget(self.frictionRange1, 1, 1, 1, 1)
        lastCalibVbox.addWidget(self.endLabel2, 2, 0, 1, 1)
        lastCalibVbox.addWidget(self.frictionRange2, 2, 1, 1, 1)
        lastCalibVbox.addWidget(self.prl2Range1, 1, 2, 1, 1)
        lastCalibVbox.addWidget(self.prl2Range2, 2, 2, 1, 1)
        lastCalibVbox.addWidget(self.zero2Range1, 1, 3, 1, 1)
        lastCalibVbox.addWidget(self.zero2Range2, 2, 3, 1, 1)
        lastCalibVbox.addWidget(self.latCTlabel, 3, 0, 1, 1)
        lastCalibVbox.addWidget(self.latCrossTalk, 3, 1, 1, 1)

        fittingVbox = QGridLayout()
        self.fittingGroupBox.setLayout(fittingVbox)
        fittingVbox.addWidget(self.startFitLabel, 0, 0, 1, 1)
        fittingVbox.addWidget(self.endFitLabel, 1, 0, 1, 1)
        fittingVbox.addWidget(self.fitStart, 0, 1, 1, 1)
        fittingVbox.addWidget(self.fitStop, 1, 1, 1, 1)
        fittingVbox.addWidget(self.xFitLabel, 0, 2, 1, 1)
        fittingVbox.addWidget(self.yFitLabel, 1, 2, 1, 1)
        fittingVbox.addWidget(self.xFit, 0, 3, 1, 1)
        fittingVbox.addWidget(self.yFit, 1, 3, 1, 1)
        fittingVbox.addWidget(self.fitPosLabel, 0, 4, 1, 1)
        fittingVbox.addWidget(self.fitPos, 0, 5, 1, 1)
        fittingVbox.addWidget(self.showFitEq, 1, 4, 1, 2)

    def filter_change(self):
        if self.filter_wind.value() %2 == 0: #make sure its odd
            self.filter_wind.blockSignals(True)
            self.filter_wind.setValue(self.filter_wind.value() + 1)
            self.filter_wind.blockSignals(False)

    def update_range(self):
        key = self.roiChoice.currentText()
        if key not in self.rangeDict.keys():
            key = "Default"

        self.zeroRange1.blockSignals(True)
        self.zeroRange1.setValue(self.rangeDict[key][0][0])
        self.zeroRange1.blockSignals(False)
        self.zeroRange2.blockSignals(True)
        self.zeroRange2.setValue(self.rangeDict[key][0][1])
        self.zeroRange2.blockSignals(False)
        self.adhRange1.blockSignals(True)
        self.adhRange1.setValue(self.rangeDict[key][1][0])
        self.adhRange1.blockSignals(False)
        self.adhRange2.blockSignals(True)
        self.adhRange2.setValue(self.rangeDict[key][1][1])
        self.adhRange2.blockSignals(False)
        self.prl1Range1.blockSignals(True)
        self.prl1Range1.setValue(self.rangeDict[key][2][0])
        self.prl1Range1.blockSignals(False)
        self.prl1Range2.blockSignals(True)
        self.prl1Range2.setValue(self.rangeDict[key][2][1])
        self.prl1Range2.blockSignals(False)
        self.frictionRange1.blockSignals(True)
        self.frictionRange1.setValue(self.rangeDict[key][3][0])
        self.frictionRange1.blockSignals(False)
        self.frictionRange2.blockSignals(True)
        self.frictionRange2.setValue(self.rangeDict[key][3][1])
        self.frictionRange2.blockSignals(False)
        self.prl2Range1.blockSignals(True)
        self.prl2Range1.setValue(self.rangeDict[key][4][0])
        self.prl2Range1.blockSignals(False)
        self.prl2Range2.blockSignals(True)
        self.prl2Range2.setValue(self.rangeDict[key][4][1])
        self.prl2Range2.blockSignals(False)
        self.zero2Range1.blockSignals(True)
        self.zero2Range1.setValue(self.rangeDict[key][5][0])
        self.zero2Range1.blockSignals(False)
        self.zero2Range2.blockSignals(True)
        self.zero2Range2.setValue(self.rangeDict[key][5][1])
        self.zero2Range2.blockSignals(False)

    def update_dict(self):
        self.rangeDict[self.roiChoice.currentText()] = [[self.zeroRange1.value(),
                                                           self.zeroRange2.value()],
                                                        [self.adhRange1.value(),
                                                           self.adhRange2.value()],
                                                        [self.prl1Range1.value(),
                                                           self.prl1Range2.value()],
                                                        [self.frictionRange1.value(),
                                                           self.frictionRange2.value()],
                                                        [self.prl2Range1.value(),
                                                         self.prl2Range2.value()],
                                                        [self.zero2Range1.value(),
                                                           self.zero2Range2.value()]]
    def show_window(self): #show window
        self.update_dict()
        self.show()

# %% Live Plot Window  
class PlotWindow(QWidget):
    def __init__(self, parent=None):
        super(PlotWindow, self).__init__(parent=parent)

        self.setGeometry(1410, 30, 500, 400)
        self.setWindowTitle("Live Plot")
        self.resizeEvent = self.onResize

        self.layout = QGridLayout(self)

        series = QScatterSeries()
        series.append(1, 3)
        series.append(4, 5)
        series.append(5, 4.5)
        series.append(7, 1)
        series.append(11, 2)
        
        self.liveChart = QChart() #live chart
        self.liveChart.legend().hide()
        self.liveChart.setSizePolicy(QSizePolicy.Minimum, QSizePolicy.Minimum)
        self.liveChart.addSeries(series)
        self.liveChart.createDefaultAxes()      
        
        self.liveChartScene = QGraphicsScene(self)
        self.liveChartScene.addItem(self.liveChart)
        
        self.liveChartView = QGraphicsView(self.liveChartScene, self)
        self.liveChartView.setSizePolicy(QSizePolicy.Minimum, QSizePolicy.Minimum)
        self.liveChartView.setRenderHint(QPainter.Antialiasing)
        self.liveChartView.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.liveChartView.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.liveChartView.setViewportUpdateMode(QGraphicsView.MinimalViewportUpdate)

        self.layout.addWidget(self.liveChartView, 0, 0)


    def home(self):
        self.show()

    def onResize(self, event):
        self.liveChartScene.removeItem(self.liveChart)
        self.liveChartView.resetCachedContent()
        w, h = (self.liveChartView.size().width(),self.liveChartView.size().height())

        self.liveChart.setMinimumSize(w, h)
        self.liveChart.createDefaultAxes()
        self.liveChartScene.setSceneRect(0, 0, w, h)
        self.liveChartScene.addItem(self.liveChart)
        self.liveChartView.setScene(self.liveChartScene)

# %% Zoomable QGraphicsView Display
class MyQGraphicsView(QGraphicsView): #zoom QFraphicsView
    def __init__ (self, parent=None):
        super(MyQGraphicsView, self).__init__ (parent)

    def wheelEvent(self, event):
       
        zoomInFactor = 1.25  # Zoom Factor
        zoomOutFactor = 1 / zoomInFactor

        # Set Anchors
        self.setTransformationAnchor(QGraphicsView.NoAnchor)
        self.setResizeAnchor(QGraphicsView.NoAnchor)

        # Save the scene pos
        oldPos = self.mapToScene(event.pos())

        # Zoom
        if event.angleDelta().y() > 0:
            zoomFactor = zoomInFactor
        else:
            zoomFactor = zoomOutFactor
        self.scale(zoomFactor, zoomFactor)

        # Get the new position
        newPos = self.mapToScene(event.pos())

        # Move scene to old position
        delta = newPos - oldPos
        self.translate(delta.x(), delta.y())

# %% Video frame count thread
class CountFrameThread(QThread):
    output = pyqtSignal('PyQt_PyObject')
    def __init__(self, cap):
        QThread.__init__(self)
        self.cap = cap

    def __del__(self):
        self.wait()

    def run(self):
        i = 1
        while(True):
            ret, frame = self.cap.read()
##            self.emit(SIGNAL('frame_number'), i)
            self.output.emit(i)
            i += 1
            if ret ==False:
                self.frameCount = self.cap.get(cv2.CAP_PROP_POS_FRAMES)
                print("framecount", self.frameCount)
##                self.cap.release()
                break
# %% Save contour data thread
class ContourDataThread(QThread): 
    output = pyqtSignal('PyQt_PyObject')
    def __init__(self, df_contour, contourDataPath, roi_labels):
        QThread.__init__(self)
        self.df_contour = df_contour
        self.contourDataPath = contourDataPath
        self.roi_labels = roi_labels

    def __del__(self):
        self.wait()

    def run(self):
        #delete old data in appended file
        self.output.emit("Creating contour dataset..")
        for r in self.df_contour['ROI_Label'].unique():
            if r not in self.roi_labels:
                self.df_contour = self.df_contour[self.df_contour['ROI_Label'] != r]
                continue
            for i in self.df_contour['Frame_no'].unique():
                ind_all = self.df_contour.index[(self.df_contour['Frame_no'] == i) &
                                           (self.df_contour['ROI_Label'] == r)].tolist()
                ind_0 = self.df_contour.index[(self.df_contour['Frame_no'] == i) &
                                         (self.df_contour['ROI_Label'] == r) &
                                         (self.df_contour['Contour_ID'] == 0)].tolist()
                ind_old = [a for a in ind_all if a < max(ind_0)]
                self.df_contour.drop(ind_old, inplace = True)
        self.output.emit("Saving contour data to excel..")
        self.df_contour.to_excel(self.contourDataPath) #save 
        del self.df_contour

# %% Save summary plots thread
class SummPlotThread(QThread): 
    output = pyqtSignal('PyQt_PyObject')
    def __init__(self, summary, pltformat):
        QThread.__init__(self)
        self.summary = summary
        self.pltformat = pltformat

    def __del__(self):
        self.wait()

    def run(self):
        self.output.emit("Saving summary plots..")
        self.summary.saveSummaryPlot(self.pltformat)

# %% Main Application Call
if __name__ == "__main__":
    def except_hook(cls, exception, traceback): #display error message/print traceback
    ##    print(dir(traceback))   
        root = Tk()
        root.withdraw()
        messagebox.showinfo(cls.__name__, str(exception) + '\n\nTraceback: ' +
                            str(traceback.tb_frame))
        root.destroy()    
        sys.__excepthook__(cls, exception, traceback)
             
    def run():
        app = QApplication(sys.argv)
        app.setWindowIcon(QIcon('images/icon.ico'))
        file = QFile("style/dark.qss")
        file.open(QFile.ReadOnly | QFile.Text)
        stream = QTextStream(file)
        app.setStyleSheet(stream.readAll())
        Gui = MainWindow()
        Gui.setWindowIcon(QIcon('images/icon.ico'))
        # sys.exit(app.exec_())
        Gui.show()
        app.exec_()
    
    sys.excepthook = except_hook
    run()

